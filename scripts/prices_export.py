# -*- coding: utf-8 -*-
# Выгрузка цен и ассортимента конфигуратора в Excel (задание «скрипт для цен 18,07»).
# Читает data/materials.json → пишет «для работы/цены.xlsx» (по листу на категорию).
# Правки вносятся в Excel, обратно — скриптом prices_import.py («Загрузить цены.bat»).
#
# Устройство файла:
#  - служебные колонки (_id и т.п.) скрыты — по ним импорт находит позицию, руками не трогать;
#  - листы «только цены» (Направляющие, Сетчатые полки, Корзины, Фурнитура) — менять можно
#    только цену, новые строки запрещены (ассортимент там жёстко связан с кодом конфигуратора);
#  - листы с ассортиментом (ЛДСП, Профили, Цвета профилей, Наполнение дверей, Услуги) — новая
#    строка внизу без служебного id = новая позиция;
#  - первый лист «Справка» — те же правила словами, для пользователя.
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'data', 'materials.json')
OUT_DIR = os.path.join(ROOT, 'для работы')
OUT = os.path.join(OUT_DIR, 'цены.xlsx')

SLIDE_TYPES = {'ball': 'Шариковые', 'soft': 'С доводчиком', 'push': 'Push-to-open', 'blum': 'BLUM'}
LDSP_GROUPS = [('korpus', 'ЛДСП корпус'), ('fasad', 'ЛДСП фасад'), ('fill', 'ЛДСП наполнение')]

HELP_TEXT = [
    'Как пользоваться этим файлом',
    '',
    '1. Меняйте цены прямо в жёлтых колонках — числа, без «₽» и пробелов.',
    '2. Листы с ассортиментом (ЛДСП, Профили купе, Цвета профилей, Наполнение дверей, Услуги):',
    '   новая строка внизу таблицы = новая позиция. Заполните все видимые колонки строки.',
    '   На листах ЛДСП новый производитель создаётся сам — просто напишите его имя в строке.',
    '   «Профили купе» — прайс по сочетаниям: на КАЖДУЮ пару профиль+цвет своя строка с ценами.',
    '   Новый цвет профиля: сначала строка на листе «Цвета профилей», затем строки с ценами',
    '   этого цвета для каждого профиля на листе «Профили купе» (загрузка подскажет, каких нет).',
    '   Новый профиль: просто строки с его именем и ценами на «Профили купе» (по всем цветам).',
    '3. Цвета выбираются ПО ИМЕНИ из листа «Палитра» (в ячейке цвета есть выпадающий список).',
    '   Нужен новый цвет — сначала добавьте его строкой на «Палитру» (имя + hex вида #f4f3f0),',
    '   затем выбирайте по имени. Можно вписать hex и напрямую, если так удобнее.',
    '4. У ЛДСП цвет не задаётся вовсе: настоящий вид даст текстура («Файл текстуры» — имя jpg',
    '   из папки data/textures). Пока текстуры нет, плитка в конфигураторе коричневая — это',
    '   индикатор «текстура не загружена», так и задумано.',
    '5. Листы только с ценами (Направляющие, Сетчатые полки, Корзины, Фурнитура и разное):',
    '   новые строки добавлять нельзя — только менять цены. Серые строки-заголовки внутри',
    '   листа просто разделяют блоки с разными единицами измерения.',
    '6. Скрытые колонки (_id и похожие) не трогать — по ним загрузка находит позиции.',
    '7. Удалять строки нельзя — вывод позиции из ассортимента будет отдельным скриптом.',
    '8. Названия листов и порядок колонок менять нельзя — загрузка ищет данные именно по ним.',
    '   Правьте файл, который сделала выгрузка, а не собирайте свой с нуля.',
    '9. Когда закончили — сохраните файл и запустите «Загрузить цены.bat».',
    '   Загрузка сначала всё проверит: при любой ошибке ничего не изменится, будет список ошибок.',
]


def main():
    global OUT
    # Путь можно передать аргументом — для автоматизации/проверок
    if len(sys.argv) > 1:
        OUT = sys.argv[1]
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    with open(SRC, encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()
    bold = Font(bold=True)
    price_fill = PatternFill('solid', fgColor='FFF6D5')   # жёлтый — редактируемые цены
    section_fill = PatternFill('solid', fgColor='E4E4E8')  # серый — строки-разделители блоков

    palette = data.get('palette', [])
    pal_by_hex = {p['hex'].lower(): p['name'] for p in palette}

    def sheet(title, headers, hidden_cols=(), price_cols=(), widths=()):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = bold
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        for c in hidden_cols:
            ws.column_dimensions[get_column_letter(c)].hidden = True
        ws.freeze_panes = 'A2'
        ws._price_cols = price_cols
        return ws

    def add_row(ws, values):
        ws.append(values)
        r = ws.max_row
        for c in ws._price_cols:
            ws.cell(row=r, column=c).fill = price_fill

    def add_section(ws, title, ncols):
        # Серый разделитель блока (например, разные единицы измерения на одном листе);
        # загрузка такие строки пропускает — у них нет ни цены, ни служебного ключа
        ws.append([title])
        r = ws.max_row
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = section_fill
        ws.cell(row=r, column=1).font = bold

    def color_cell(ws, col_idx, hexv, dv):
        """Ячейка цвета: имя из палитры (если есть), закрашена своим цветом, с выпадающим списком."""
        cell = ws.cell(row=ws.max_row, column=col_idx)
        hx = (hexv or '').lower()
        cell.value = pal_by_hex.get(hx, hexv)
        if hx:
            cell.fill = PatternFill('solid', fgColor=hx.lstrip('#'))
        dv.add(cell)

    # Справка
    ws = wb.active
    ws.title = 'Справка'
    for line in HELP_TEXT:
        ws.append([line])
    ws.column_dimensions['A'].width = 100
    ws['A1'].font = bold

    # Палитра — служебный справочник «имя → hex» (просьба 21.07): цвета в остальных листах
    # выбираются по имени отсюда; новая строка = новый цвет в палитре
    ws = sheet('Палитра', ['Название', 'Цвет (hex)'], price_cols=(), widths=(22, 12))
    for p in palette:
        add_row(ws, [p['name'], p['hex']])
        ws.cell(row=ws.max_row, column=2).fill = PatternFill('solid', fgColor=p['hex'].lstrip('#'))
    # выпадающий список для ячеек цвета на других листах (диапазон с запасом на новые строки)
    def make_color_dv():
        dv = DataValidation(type='list', formula1='=Палитра!$A$2:$A$200', allow_blank=True,
                            showErrorMessage=False)
        return dv

    # ЛДСП: корпус / фасад / наполнение — без цвета (вид задаст текстура, без неё — коричневый)
    for group, title in LDSP_GROUPS:
        ws = sheet(title,
                   ['Производитель', 'Название цвета', 'Цена ₽/м²', 'Файл текстуры', '_id', '_producer'],
                   hidden_cols=(5, 6), price_cols=(3,), widths=(18, 26, 12, 20, 10, 12))
        for prod in data[group]['producers']:
            for col in prod['colors']:
                add_row(ws, [prod['name'], col['name'], col['pricePerM2'],
                             col.get('texture', ''), col['id'], prod['id']])

    # Профили купе — полный прайс по сочетаниям профиль×цвет (не коэффициенты, просьба 21.07)
    ws = sheet('Профили купе',
               ['Профиль', 'Цвет', 'Вертикаль ₽/пог.м', 'Горизонт. верх ₽/пог.м', 'Горизонт. низ ₽/пог.м', '_key'],
               hidden_cols=(6,), price_cols=(3, 4, 5), widths=(18, 14, 18, 20, 20, 16))
    prof_names = {p['id']: p['name'] for p in data['slidingDoor']['profiles']}
    col_names = {c['id']: c['name'] for c in data['slidingDoor']['colors']}
    for pp in data['slidingDoor'].get('profilePrices', []):
        add_row(ws, [prof_names.get(pp['profile'], pp['profile']), col_names.get(pp['color'], pp['color']),
                     pp['vertPerM'], pp['horizTopPerM'], pp['horizBottomPerM'],
                     f"{pp['profile']}:{pp['color']}"])

    # Цвета профилей (только ассортимент — цены на листе «Профили купе»); цвет — по имени
    # из «Палитры», ячейка закрашена и с выпадающим списком
    ws = sheet('Цвета профилей',
               ['Название', 'Цвет (из палитры)', '_id'],
               hidden_cols=(3,), price_cols=(), widths=(20, 18, 10))
    dv = make_color_dv()
    ws.add_data_validation(dv)
    for c in data['slidingDoor']['colors']:
        add_row(ws, [c['name'], None, c['id']])
        color_cell(ws, 2, c['hex'], dv)

    # Наполнение дверей: зеркало (только цена) + стёкла (ассортимент)
    ws = sheet('Наполнение дверей',
               ['Тип', 'Название', 'Цвет (из палитры)', 'Цена ₽/м²', '_id'],
               hidden_cols=(5,), price_cols=(4,), widths=(12, 22, 18, 12, 10))
    dv = make_color_dv()
    ws.add_data_validation(dv)
    fills = data['slidingDoor']['fills']
    add_row(ws, ['Зеркало', fills['mirror'].get('name', 'Зеркало'), '', fills['mirror']['pricePerM2'], 'mirror'])
    for c in fills.get('glass', {}).get('colors', []):
        add_row(ws, ['Стекло', c['name'], None, c['pricePerM2'], c['id']])
        color_cell(ws, 3, c['color'], dv)

    # Направляющие (только цены, размерная сетка фиксирована)
    ws = sheet('Направляющие',
               ['Тип', 'Длина, мм', 'Цена ₽/компл.', '_key'],
               hidden_cols=(4,), price_cols=(3,), widths=(16, 12, 14, 14))
    for s in data['drawerSlide']:
        add_row(ws, [SLIDE_TYPES.get(s['type'], s['type']), s['length'], s['price'], f"{s['type']}:{s['length']}"])

    # Сетчатые полки (только цены)
    ws = sheet('Сетчатые полки',
               ['Название', 'Цена ₽/пог.м', '_key'],
               hidden_cols=(3,), price_cols=(2,), widths=(22, 14, 14))
    for m in data['meshShelf']:
        add_row(ws, [m['name'], m['pricePerM'], f"{m['depth']}:{m['color']}"])

    # Корзины (только цены)
    ws = sheet('Корзины',
               ['Ширина', 'Глубина', 'Высота', 'Цвет', 'Цена ₽', '_key'],
               hidden_cols=(6,), price_cols=(5,), widths=(10, 10, 10, 12, 12, 18))
    for b in data['basket']:
        add_row(ws, [b['width'], b['depth'], b['height'], b['color'], b['price'],
                     f"{b['width']}:{b['depth']}:{b['height']}:{b['color']}"])

    # Фурнитура и разное (только цены): блоки по единицам измерения с серыми
    # строками-разделителями (просьба 21.07 — не смешивать разные единицы в одной таблице)
    ws = sheet('Фурнитура и разное',
               ['Позиция', 'Цена ₽', '_key'],
               hidden_cols=(3,), price_cols=(2,), widths=(44, 12, 20))
    sd = data['slidingDoor']
    add_section(ws, '— За штуку / комплект —', 2)
    for it in data['fittings']:
        add_row(ws, [it['name'], it['price'], f"fittings:{it['id']}"])
    add_row(ws, [data['swingDoorHardware']['name'], data['swingDoorHardware']['pricePerDoor'], 'swing'])
    add_row(ws, [sd['rollers']['name'], sd['rollers']['pricePerSet'], 'rollers'])
    add_section(ws, '— За погонный метр —', 2)
    add_row(ws, [sd['track']['name'], sd['track']['pricePerM'], 'track'])
    add_row(ws, [sd['divider']['name'], sd['divider']['pricePerM'], 'divider'])
    add_row(ws, ['Кромка', data['edgeBanding']['pricePerM'], 'edge'])

    # Услуги (extras) — ассортимент расширяемый
    ws = sheet('Услуги',
               ['Группа', 'Название', 'Цена ₽', '_group', '_id'],
               hidden_cols=(4, 5), price_cols=(3,), widths=(20, 40, 12, 12, 12))
    for grp in data['extras']:
        for it in grp['items']:
            add_row(ws, [grp['name'], it['name'], it['price'], grp['id'], it['id']])

    os.makedirs(OUT_DIR, exist_ok=True)
    try:
        wb.save(OUT)
    except PermissionError:
        print('ОШИБКА: файл «цены.xlsx» открыт в Excel — закройте его и запустите выгрузку ещё раз.')
        return 1
    print(f'Готово: {OUT}')
    print('Листы: ' + ', '.join(wb.sheetnames))
    print('Правьте цены/ассортимент и запускайте «Загрузить цены.bat».')
    return 0


if __name__ == '__main__':
    code = main()
    input('\nНажмите Enter, чтобы закрыть...')
    sys.exit(code)
