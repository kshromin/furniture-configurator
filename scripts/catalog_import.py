# -*- coding: utf-8 -*-
# Загрузка правок ассортимента/цен из Excel обратно в data/materials.json (Этап 2 плана «первым
# делом»). Пара к catalog_export.py. Модель: находит позицию по скрытому `_key`, применяет цену
# (и доп. поля). БЕЗОПАСНО: сначала всё проверяет на КОПИИ; при любой ошибке файл не меняется —
# печатается список ошибок. Перед записью делает backup materials.json.bak.
#
# Запуск:
#   python catalog_import.py                 → окно выбора xlsx
#   python catalog_import.py --in <path>     → без окна (для тестов/автоматизации)
#
# КЛЮЧ = СТАБИЛЬНЫЙ ID (задание «формат выгрузки»): ключ тот же — значит та же позиция, и правка
# ЛЮБОГО столбца (название, цвет, размер, производитель, цена) обновляет её, а не заводит дубль;
# строка без ключа = новая позиция. Позиции, у которых ключ раньше собирался из атрибутов (ЛДСП,
# сетки, корзины, направляющие), опознаются по `gid` — см. ensure_gids(); старые ключи понимаются
# как раньше. Названия, которых нет в базе (направляющие, общие элементы профиля), хранятся в
# data['catalogLabels'] — см. set_label().
#
# Умеет: (1) обновлять цены/поля существующих позиций по `_key`; (2) создавать НОВЫЕ позиции из
# строк без `_key` на расширяемых листах (ЛДСП — с автосозданием кромок 16/32, Наполнение дверей —
# стекло, Доп.элементы, Сетчатые полки, Корзины). На фиксированных листах (профили, направляющие,
# фурнитура) новые строки пропускаются. id новых позиций — транслит имени (slugify).
import json
import os
import sys
import copy
import re

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DST = os.path.join(ROOT, 'data', 'materials.json')
IN_DIR = os.path.normpath(os.path.join(ROOT, '..', 'Выгрузки'))
MANUAL = 'вручную'

# ЕДИНЫЙ формат колонок — одинаковый на всех листах (см. HEADERS в catalog_export.py), 1-based.
COLS = dict(key=1, producer=2, name=3, color=4, unit=5, price=6, h=7, l=8, w=9,
            dep=10, hex=11, korpus=12, fasad=13, fill=14, texture=15)
SHEET_NAMES = ['МАТЕРИАЛ', 'ЛДСП', 'Кромка', 'Наполнение дверей', 'Профили купе', 'Цвета профилей',
               'Сетчатые полки', 'Корзины', 'Направляющие', 'Фурнитура', 'Услуги', 'Доп.элементы']
# Поля, которые следуют из самой позиции (в catalogMeta не пишем) — см. DERIVED в catalog_export.py.
try:
    from catalog_export import DERIVED, SLIDE_TYPES, ELEMENT_LABELS, DEFAULT_SERVICE_GROUPS
except Exception:
    DEFAULT_SERVICE_GROUPS = {'delivery', 'lift', 'montage'}
    DERIVED = {'ldsp': {'h', 'hex', 'dep', 'producer', 'l', 'w'},
               'ldspm': {'h', 'hex'}, 'edge': {'h', 'dep'}, 'dfill': {'hex'}, 'prof': {'hex'},
               'profcol': {'hex'}, 'mesh': {'w'}, 'basket': {'h', 'l', 'w'}, 'slide': {'l'}}
    SLIDE_TYPES = {'ball': 'Шариковые', 'soft': 'С доводчиком', 'push': 'Push-to-open', 'blum': 'BLUM'}
    ELEMENT_LABELS = {'horizTop': 'Горизонт верхний', 'horizBottom': 'Горизонт нижний',
                      'divider': 'Перемычка', 'track': 'Направляющая (верх+низ)'}
# Списки, позиции которых опознаются по стабильному `gid` (ключ = gid): правка размера/цвета меняет
# ТУ ЖЕ позицию. gid выводится из тех же полей, из которых раньше собирался ключ, — старые выгрузки
# грузятся без потерь. Значение — поля gid + поля, по которым позиция должна остаться уникальной.
GID_FIELDS = {'meshShelf': ('depth', 'color'), 'basket': ('width', 'depth', 'height', 'color'),
              'drawerSlide': ('type', 'length')}
# Свободные колонки: то, что приложение по этой категории не использует, но пользователь вписал —
# храним в catalogMeta и возвращаем в следующую выгрузку (заполнять базу можно целиком, а как это
# показывать в конфигураторе — решается позже). Что из них категория ПРИМЕНЯЕТ — см. DERIVED.
META_FIELDS = ('producer', 'h', 'l', 'w', 'dep', 'hex', 'unit', 'color', 'texture')
# Ед.изм печатает сама выгрузка, поэтому запоминаем её, только если пользователь её изменил
# (иначе в файл осела бы «кв.м» у каждой строки).
META_ONLY_IF_CHANGED = ('unit',)
# Позиции, заведённые в самой программе (верхняя таблица листов «Фурнитура»/«Услуги»): их нельзя
# добавлять/удалять, поэтому у них правится ТОЛЬКО цена — исключение из правила «правь любой
# столбец» (прямое указание пользователя в задании «формат выгрузки»).
FIXED_PRICE_ONLY = {'fit', 'swing', 'rollers', 'rod', 'softclose', 'handle', 'service'}
UNITS_FIT = ('шт', 'комплект', 'пог.м')  # других единиц у фурнитуры нет (задание, п.8)
# ключи-шаблоны (ручные/справочные) — не позиции, пропускаем при записи
TEMPLATE_KEYS = {'dfill:special', 'addon:custom:manual'}

REV_SURFACE = {label: key for key, label in [('korpus', 'корпус'), ('fasad', 'фасад'), ('fill', 'наполнение')]}
REV_METAL = {'белый': 'white', 'серебро': 'silver', 'чёрный': 'black', 'черный': 'black'}

_TR = {'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e', 'ж': 'zh', 'з': 'z',
       'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
       'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'h', 'ц': 'c', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
       'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya'}


def slugify(name, existing, prefix='item'):
    """Стабильный ascii-id из имени (транслит), уникальный среди existing."""
    s = ''.join(_TR.get(ch, ch if (ch.isascii() and ch.isalnum()) else ' ') for ch in str(name).lower())
    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')[:24] or prefix
    base, i = s, 2
    while s in existing:
        s = f'{base}_{i}'; i += 1
    return s


# ── Создатели НОВЫХ позиций (строка без _key) по имени листа ─────────────────────────────────
def new_ldsp(data, e, ctx):
    # Единый формат: Производитель | Название(«ЛДСП») | Цвет | Ед.изм | Цена | Высота(=толщина) |
    # … | Корпус | Фасад | Наполнение (да/нет) | Файл текстуры. Материал заводится во все «да».
    pname = str(e.get('producer') or '').strip()
    disp = str(e.get('color') or '').strip()
    if not disp:
        return f'{ctx}: пустой цвет (колонка «Цвет»)'
    if not pname:
        return f'{ctx}: пустой производитель'
    cname = disp                      # «Цвет» — имя материала; «Название» — тип (отдельно)
    th = int(num(e.get('h')) or 16)
    p = num(e.get('price'))
    if p is None:
        return f'{ctx}: цена не число'
    surfaces = [('korpus', e.get('korpus')), ('fasad', e.get('fasad')), ('fill', e.get('fill'))]
    if not any(is_yes(v) for _, v in surfaces):
        return f'{ctx}: не отмечена ни одна поверхность (Корпус/Фасад/Наполнение)'
    maxp = max_part(e, ctx)
    if isinstance(maxp, str):
        return maxp
    existing = {c.get('gid') for s in ('korpus', 'fasad', 'fill')
                for prod in data[s]['producers'] for c in prod['colors'] if c.get('gid')}
    gid = slugify(f'{pname}_{cname}_{th}', existing, 'ldsp')  # новый стабильный gid
    for surf, flag in surfaces:
        if is_yes(flag):
            create_ldsp_member(data, surf, gid, pname, cname, th, p, e.get('texture'), e.get('hex'), maxp, e.get('name'))
    return None


def new_dfill(data, e, ctx):
    """Новая строка на листе «Наполнение дверей»: «Название» = ТИП (Зеркало, Стекло, Ротанг…),
    «Цвет» = позиция этого типа. Знакомый тип — позиция добавится в него, новый — заведётся тип
    (в fills['extra']). Конфигуратор сегодня показывает ЛДСП/зеркало/стекло/спеццвет, остальные
    типы копятся в каталоге до того, как сделаем их выбор в интерфейсе."""
    p = num(e.get('price'))
    if p is None:
        return f'{ctx}: цена не число'
    cname = txt(e.get('color'))
    if not cname:
        return f'{ctx}: пустое название позиции (колонка «Цвет»)'
    tname = txt(e.get('name'))
    if not tname:
        return f'{ctx}: не указан тип наполнения (колонка «Название», напр. Стекло)'
    fills = data['slidingDoor']['fills']
    hexv = txt(e.get('hex'))
    if tname.lower() in ('стекло', str(fills.get('glass', {}).get('name', '')).strip().lower()):
        cols = fills['glass']['colors']       # стекло — тип, который приложение уже умеет
        cid = slugify(cname, {c['id'] for c in cols}, 'glass')
        cols.append({'id': cid, 'name': cname, 'color': hexv or '#d9ecf0', 'pricePerM2': p})
        return None
    extra = fills.setdefault('extra', [])
    t = next((z for z in extra if z['name'].strip().lower() == tname.lower()), None)
    if t is None:
        tid = slugify(tname, {z['id'] for z in extra} | {'mirror', 'glass'}, 'fill')
        t = {'id': tid, 'name': tname, 'colors': []}
        extra.append(t)
    cid = slugify(cname, {c['id'] for c in t['colors']}, 'fill')
    t['colors'].append({'id': cid, 'name': cname, 'color': hexv or '', 'pricePerM2': p})
    return None


def new_addon_item(data, e, ctx, kind):
    """Новая строка в добавляемой таблице «Услуг» / «Доп.элементов»: «Название» = позиция,
    «От чего зависит» = раздел (нет такого — заводится новый, вид определяет лист)."""
    price = e.get('price')
    manual = isinstance(price, str) and price.strip().lower() == MANUAL
    p = MANUAL if manual else num(price)
    if p is None:
        return f'{ctx}: цена не число'
    item_name = txt(e.get('name'))
    if not item_name:
        return f'{ctx}: пустое название позиции (колонка «Название»)'
    grp_name = txt(e.get('dep'))
    if not grp_name:
        return f'{ctx}: не указан раздел (колонка «От чего зависит»)'
    g = next((x for x in data['extras'] if x['name'] == grp_name), None)
    if g is None:
        gid = slugify(grp_name, {x['id'] for x in data['extras']}, 'grp')
        g = {'id': gid, 'name': grp_name, 'kind': kind, 'items': []}
        data['extras'].append(g)
    iid = slugify(item_name, {it['id'] for it in g['items']}, 'addon')
    item = {'id': iid, 'name': item_name, 'price': 0 if manual else p}
    if manual:
        item['manual'] = True
    g['items'].append(item)
    return None


def fit_unit(e):
    """Ед.изм добавляемой фурнитуры: только шт / комплект / пог.м (None — если что-то другое)."""
    u = txt(e.get('unit')) or UNITS_FIT[0]
    return u if u in UNITS_FIT else None


def new_fitopt(data, e, ctx):
    """Новая строка в добавляемой таблице «Фурнитуры»: позиция ассортимента в разделе
    («Ручки ящика» и т.п.). Приложение пока выбирает фурнитуру не по разделам — это каталог."""
    p = num(e.get('price'))
    if p is None:
        return f'{ctx}: цена не число'
    name = txt(e.get('name'))
    if not name:
        return f'{ctx}: пустое название позиции (колонка «Название»)'
    grp = txt(e.get('dep'))
    if not grp:
        return f'{ctx}: не указан раздел (колонка «От чего зависит», напр. «Ручки ящика»)'
    unit = fit_unit(e)
    if unit is None:
        return f'{ctx}: ед.изм «{txt(e.get("unit"))}» — допустимы только ' + ', '.join(UNITS_FIT)
    lst = data.setdefault('fittingOptions', [])
    gid = slugify(name, {i['gid'] for i in lst}, 'fit')
    lst.append({'gid': gid, 'group': grp, 'name': name, 'unit': unit, 'price': p})
    return None


def new_mesh(data, e, ctx):
    d, p = num(e.get('w')), num(e.get('price'))  # глубина полки — колонка «Ширина, мм»
    if d is None or p is None:
        return f'{ctx}: глубина (ширина, мм) / цена не число'
    color = REV_METAL.get(str(e.get('color') or '').strip().lower(), e.get('color'))
    data['meshShelf'].append({'depth': int(d), 'color': color,
                              'name': str(e.get('name') or 'Сетчатая полка'), 'pricePerM': p})
    return None


def new_basket(data, e, ctx):
    # Высота | Длинна(=ширина корзины) | Ширина(=глубина корзины) — как в выгрузке.
    n = [num(e.get('l')), num(e.get('w')), num(e.get('h')), num(e.get('price'))]
    if any(x is None for x in n):
        return f'{ctx}: размеры/цена не число'
    color = REV_METAL.get(str(e.get('color') or '').strip().lower(), e.get('color'))
    data['basket'].append({'width': int(n[0]), 'depth': int(n[1]), 'height': int(n[2]),
                           'color': color, 'price': n[3]})
    return None


def store_meta(data, key, e, base=None):
    """Свободные поля (производитель/габариты/зависимость/hex/ед.изм/цвет/текстура), которых нет в
    самой базе по этой категории, храним в data['catalogMeta'][key] — так они переживают выгрузку/
    загрузку. Поля, которые категория ПРИМЕНЯЕТ или выводит из позиции (DERIVED), не пишем: они и
    так вернутся из базы. Пустая ячейка = значение забыли."""
    derived = DERIVED.get(str(key).split(':')[0], set())
    meta = data.setdefault('catalogMeta', {})
    vals = dict(meta.get(key, {}))
    for f in META_FIELDS:
        if f in derived:
            vals.pop(f, None)
            continue
        v = e.get(f)
        v = v.strip() if isinstance(v, str) else v
        if v in (None, ''):
            vals.pop(f, None)
            continue
        # значение, которое печатает сама выгрузка, запоминаем только если его изменили
        if f in META_ONLY_IF_CHANGED and f not in vals and base is not None \
                and str(v) == str(base.get(f) or '').strip():
            continue
        vals[f] = v
    if vals:
        meta[key] = vals
    else:
        meta.pop(key, None)


CREATORS = {'МАТЕРИАЛ': new_ldsp, 'ЛДСП': new_ldsp, 'Наполнение дверей': new_dfill,
            'Доп.элементы': lambda d, e, c: new_addon_item(d, e, c, 'extra'),
            'Услуги': lambda d, e, c: new_addon_item(d, e, c, 'service'),
            'Фурнитура': new_fitopt,
            'Сетчатые полки': new_mesh, 'Корзины': new_basket}


def num(v):
    """Число из ячейки или None, если не число."""
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip().replace(' ', '').replace(' ', '').replace(',', '.')
        try:
            f = float(s)
            return int(f) if f == int(f) else f
        except ValueError:
            return None
    return None


def find_color(data, surface, prodid, colid):
    for p in data.get(surface, {}).get('producers', []):
        if p['id'] == prodid:
            for c in p.get('colors', []):
                if c['id'] == colid:
                    return c
    return None


YES_WORDS = {'да', 'yes', '1', '+', 'true', 'x', 'х', '✓', 'v'}
def is_yes(v):
    return str(v or '').strip().lower() in YES_WORDS


def find_ldsp(data, surface, prod_name, col_name, thickness):
    """Ищет (производитель, цвет) по ЧЕЛОВЕЧЕСКИМ полям — id сохраняется (не пересоздаём). Возвращает
    (producer_dict|None, color_dict|None)."""
    prod = next((p for p in data.get(surface, {}).get('producers', []) if p['name'] == prod_name), None)
    if prod is None:
        return None, None
    col = next((c for c in prod['colors']
                if c['name'] == col_name and int(c.get('thickness', 16)) == int(thickness)), None)
    return prod, col


def upsert_ldsp(data, surface, prod_name, col_name, thickness, price, texture, hexv=None):
    """Обновить цену/текстуру существующего цвета (id сохраняется) или создать новый в этой поверхности."""
    prod, col = find_ldsp(data, surface, prod_name, col_name, thickness)
    if col is not None:
        col['pricePerM2'] = price
        if texture not in (None, ''):
            col['texture'] = str(texture)
        if hexv not in (None, ''):
            col['color'] = str(hexv)
        return
    if prod is None:
        pid = slugify(prod_name or 'prod', {p['id'] for p in data[surface]['producers']}, 'prod')
        prod = {'id': pid, 'name': str(prod_name or pid), 'colors': []}
        data[surface]['producers'].append(prod)
    cid = slugify(col_name, {c['id'] for c in prod['colors']}, 'col')
    col = {'id': cid, 'name': col_name, 'color': '', 'thickness': int(thickness),
           'pricePerM2': price, 'edgePerM16': 0, 'edgePerM32': 0}
    if texture not in (None, ''):
        col['texture'] = str(texture)
    prod['colors'].append(col)


def set_kind(col, kind):
    """Тип материала («Название» на листе МАТЕРИАЛ: ЛДСП, МДФ, Массив…) — ОТДЕЛЬНОЕ поле `kind`,
    с «Цветом» не склеивается (задание 5.08: каждому столбцу — своя запись). Пусто = убрать."""
    k = str(kind or '').strip()
    if k:
        col['kind'] = k
    else:
        col.pop('kind', None)


def ldsp_full_name(typ, disp):
    """ЛЕГАСИ (лист «ЛДСП» формата сессии 67): там тип и цвет склеивались в одно имя."""
    disp = str(disp or '').strip()
    if str(typ or '').strip().lower() == 'лдсп' and not disp.lower().startswith('лдсп'):
        return 'ЛДСП ' + disp
    return disp


def ensure_ldsp_gids(data):
    """Присвоить каждому материалу ЛДСП общий стабильный gid (группа по производитель+имя+толщина
    через все поверхности). Детерминированно = id первого встреченного члена (korpus→fasad→fill),
    поэтому совпадает с тем, что выводит экспорт до первой загрузки. Идемпотентно."""
    order = ('korpus', 'fasad', 'fill')
    gids = {}  # (произв., имя, толщина) -> gid
    for surf in order:  # сперва подхватить уже проставленные gid
        for prod in data.get(surf, {}).get('producers', []):
            for c in prod['colors']:
                k = (prod['name'], c['name'], int(c.get('thickness', 16)))
                if c.get('gid') and k not in gids:
                    gids[k] = c['gid']
    for surf in order:
        for prod in data.get(surf, {}).get('producers', []):
            for c in prod['colors']:
                k = (prod['name'], c['name'], int(c.get('thickness', 16)))
                gids.setdefault(k, f"{prod['id']}__{c['id']}")
                c['gid'] = gids[k]


def ensure_gids(data):
    """Стабильные id для всех опознаваемых по gid позиций (ЛДСП + сетки/корзины/направляющие).
    Идемпотентно; вызывается до и после обработки книги (чтобы новые строки тоже получили id)."""
    ensure_ldsp_gids(data)
    for coll, fields in GID_FIELDS.items():
        items = data.get(coll) or []
        used = {it['gid'] for it in items if it.get('gid')}
        for it in items:
            if it.get('gid'):
                continue
            base = '_'.join(str(it[f]) for f in fields)
            gid, i = base, 2
            while gid in used:
                gid = f'{base}_{i}'; i += 1
            it['gid'] = gid
            used.add(gid)
    # вид группы `extras` — услуга или доп.элемент (по нему выбирается лист выгрузки)
    for grp in data.get('extras') or []:
        grp.setdefault('kind', 'service' if grp['id'] in DEFAULT_SERVICE_GROUPS else 'extra')


def by_gid(items, gid):
    return next((it for it in (items or []) if it.get('gid') == gid), None)


def clash(items, hit, fields):
    """Не совпала ли позиция после правки размеров/цвета с ДРУГОЙ такой же (дубль в каталоге)."""
    sig = tuple(str(hit[f]) for f in fields)
    return any(it is not hit and tuple(str(it[f]) for f in fields) == sig for it in items)


def txt(v):
    """Непустой текст из ячейки или None."""
    s = str(v).strip() if v is not None else ''
    return s or None


def build_baseline(original):
    """Как выглядела бы выгрузка ДО правок: ключ → значения ячеек. Нужно, чтобы отличать реально
    отредактированную ячейку от повтора того же значения в соседних строках: название элемента
    профиля и цвет профиля печатаются в КАЖДОЙ строке листа, и неотредактированные строки иначе
    затирали бы правку соседней."""
    try:
        from catalog_export import CATEGORIES
    except Exception:
        return {}
    snap = copy.deepcopy(original)
    ensure_gids(snap)  # ключи такие же, как в свежей выгрузке
    base = {}
    for _k, _label, builder in CATEGORIES:
        for r in builder(snap)['rows']:
            base[str(r[0])] = {f: (r[c - 1] if len(r) >= c else None) for f, c in COLS.items()}
    return base


def edited(extra, base, field):
    """Текст ячейки, только если он ОТЛИЧАЕТСЯ от значения в выгрузке (иначе None = не правили)."""
    v = txt(extra.get(field))
    if v is None:
        return None
    b = base.get(field)
    return None if v == (str(b).strip() if b is not None else '') else v


def set_label(data, key, name, default):
    """Название позиции, у которой в базе нет поля имени (направляющие, общие элементы профиля):
    храним переопределение в data['catalogLabels'][ключ]; совпало с типовым — не храним."""
    labels = data.setdefault('catalogLabels', {})
    name = txt(name)
    if name and name != default:
        labels[key] = name
    else:
        labels.pop(key, None)
    if not labels:
        data.pop('catalogLabels', None)


def move_to_producer(data, surf, src_prod, col, pname):
    """Правка колонки «Производитель» у ЛДСП = перенос материала к другому производителю."""
    if src_prod['name'] == pname:
        return
    dst = next((p for p in data[surf]['producers'] if p['name'] == pname), None)
    if dst is None:
        pid = slugify(pname, {p['id'] for p in data[surf]['producers']}, 'prod')
        dst = {'id': pid, 'name': pname, 'colors': []}
        data[surf]['producers'].append(dst)
    src_prod['colors'].remove(col)
    if any(c['id'] == col['id'] for c in dst['colors']):
        col['id'] = slugify(col['name'], {c['id'] for c in dst['colors']}, 'col')
    dst['colors'].append(col)
    if not src_prod['colors']:
        data[surf]['producers'].remove(src_prod)


def find_ldsp_by_gid(data, gid):
    """Все члены материала (по поверхностям) с этим gid: список (surface, prod, color)."""
    out = []
    for surf in ('korpus', 'fasad', 'fill'):
        for prod in data.get(surf, {}).get('producers', []):
            for c in prod['colors']:
                if c.get('gid') == gid:
                    out.append((surf, prod, c))
    return out


def max_part(e, errctx):
    """Макс. допустимый размер детали из материала (колонки «Длинна»/«Ширина» на листе ЛДСП):
    (длина, ширина) числом, None — если ячейка пуста (размер не задан). Текст ошибки — если не число."""
    out = []
    for fld, label in (('l', 'длинна'), ('w', 'ширина')):
        raw = e.get(fld)
        if raw in (None, ''):
            out.append(None)
            continue
        v = num(raw)
        if v is None or v <= 0:
            return f'{errctx}: макс. размер детали, {label} — не положительное число («{raw}»)'
        out.append(int(v))
    return tuple(out)


def set_max_part(col, maxp):
    """Записать/убрать макс. размер детали в самом материале (пустая ячейка = размер не задан)."""
    for fld, v in zip(('maxPartL', 'maxPartW'), maxp or (None, None)):
        if v is None:
            col.pop(fld, None)
        else:
            col[fld] = v


def create_ldsp_member(data, surf, gid, pname, cname, th, price, texture, hexv, maxp=None, kind=None):
    """Завести материал в указанной поверхности с общим gid (когда поставили «да», а члена не было)."""
    prod = next((p for p in data[surf]['producers'] if p['name'] == pname), None)
    if prod is None:
        pid = slugify(pname or 'prod', {p['id'] for p in data[surf]['producers']}, 'prod')
        prod = {'id': pid, 'name': str(pname or pid), 'colors': []}
        data[surf]['producers'].append(prod)
    cid = slugify(cname, {c['id'] for c in prod['colors']}, 'col')
    col = {'id': cid, 'gid': gid, 'name': cname, 'color': str(hexv or ''), 'thickness': int(th),
           'pricePerM2': price, 'edgePerM16': 0, 'edgePerM32': 0}
    if texture not in (None, ''):
        col['texture'] = str(texture)
    set_max_part(col, maxp)
    set_kind(col, kind)
    prod['colors'].append(col)


def apply_row(data, key, price, extra, errctx, base=None):
    """Применить одну строку к data (мутирует). Возвращает текст ошибки или None.
    `base` — значения этой строки в выгрузке ДО правок (см. build_baseline)."""
    if key in TEMPLATE_KEYS:
        return None
    base = base or {}
    parts = key.split(':')
    tag = parts[0]

    # цена (кроме листов без цены). «вручную» — числом не пишем, но ОСТАЛЬНЫЕ столбцы применяем:
    # ключ тот же = та же позиция, правка любого столбца должна дойти до базы.
    need_price = tag not in ('profcol',)
    pval, manual = None, False
    if need_price:
        if isinstance(price, str) and price.strip().lower() == MANUAL:
            manual = True
        else:
            pval = num(price)
            if pval is None:
                return f'{errctx}: цена не число («{price}»)'
            if pval < 0:
                return f'{errctx}: отрицательная цена'

    try:
        if tag == 'ldsp' and len(parts) == 2:
            # НОВЫЙ формат: ldsp:<gid> — правка по СТАБИЛЬНОМУ id. Правка ЛЮБОГО столбца (имя/цвет/
            # цена/толщина/hex/да-нет) = обновление ТОЙ ЖЕ позиции, без дублей (задание «формат
            # выгрузки»). «Цвет» — без слова «ЛДСП», возвращаем префикс в имя.
            gid = parts[1]
            pname = str(extra.get('producer') or '').strip()
            disp = str(extra.get('color') or '').strip()
            if not disp:
                return f'{errctx}: пустой цвет ЛДСП (колонка «Цвет»)'
            cname = disp                      # «Цвет» — имя материала как есть
            kind = extra.get('name')          # «Название» — ТИП материала (отдельное поле kind)
            th = int(num(extra.get('h')) or 16)
            tex, hexv = extra.get('texture'), extra.get('hex')
            maxp = max_part(extra, errctx)          # макс. допустимый размер детали (длина/ширина)
            if isinstance(maxp, str):
                return maxp
            members = find_ldsp_by_gid(data, gid)
            # 1) «нет» — убрать материал с этой поверхности (делаем ДО правок, чтобы удалять из того
            # производителя, где член лежит сейчас)
            for surf, prod, c in list(members):
                if not is_yes(extra.get(surf)):
                    prod['colors'].remove(c)
                    if not prod['colors']:
                        data[surf]['producers'].remove(prod)
                    members.remove((surf, prod, c))
            # 2) обновить ВСЕ поля оставшихся членов (правка любого столбца — та же позиция)
            for surf, prod, c in members:
                c['name'] = cname
                set_kind(c, kind)
                if pval is not None:
                    c['pricePerM2'] = pval
                c['thickness'] = th
                if hexv not in (None, ''):
                    c['color'] = str(hexv)
                if tex not in (None, ''):
                    c['texture'] = str(tex)
                set_max_part(c, maxp)
                if pname:                            # смена производителя = перенос материала
                    move_to_producer(data, surf, prod, c, pname)
            # 3) «да» там, где члена не было — завести
            if not pname and members:
                pname = members[0][1]['name']
            present = {s for s, _, _ in members}
            for surf in ('korpus', 'fasad', 'fill'):
                if is_yes(extra.get(surf)) and surf not in present:
                    create_ldsp_member(data, surf, gid, pname, cname, th, pval or 0, tex, hexv, maxp, kind)
        elif tag == 'ldspm':
            # ЛЕГАСИ (выгрузка сессии 67): одна строка = материал, поиск по имени. Оставлено для
            # загрузки прежних файлов; новые выгрузки идут форматом ldsp:<gid> выше.
            pname = str(extra.get('producer') or '').strip()
            cname = str(extra.get('color') or '').strip()
            if not pname or not cname:
                return f'{errctx}: пустой производитель/цвет ЛДСП'
            th = int(num(extra.get('h')) or 16)
            tex = extra.get('texture')
            for surf in ('korpus', 'fasad', 'fill'):
                if is_yes(extra.get(surf)):
                    upsert_ldsp(data, surf, pname, cname, th, pval, tex, extra.get('hex'))
                else:
                    prod, col = find_ldsp(data, surf, pname, cname, th)
                    if col is not None:
                        prod['colors'].remove(col)
        elif tag == 'ldsp':
            # очень старый формат (одна строка на поверхность): ldsp:surface:prodid:colid
            _, surface, prodid, colid = parts
            c = find_color(data, surface, prodid, colid)
            if c is None:
                return f'{errctx}: не найдена позиция ЛДСП {key}'
            if pval is not None:
                c['pricePerM2'] = pval
            if extra.get('texture') not in (None, ''):
                c['texture'] = str(extra['texture'])
        elif tag == 'edge':
            _, surface, prodid, colid, plate = parts
            c = find_color(data, surface, prodid, colid)
            if c is None and str(extra.get('dep') or '').startswith('ldsp:'):
                # материал мог переехать к другому производителю (правка колонки «Производитель»
                # на листе ЛДСП) — ищем по стабильному ключу из «От чего зависит»
                gid = str(extra['dep'])[5:]
                c = next((cc for s, _p, cc in find_ldsp_by_gid(data, gid) if s == surface), None)
            if c is None:
                return None  # цвет мог быть убран с этой поверхности через да/нет в листе ЛДСП —
                             # кромка для него больше не нужна (кромка хранится в самом цвете), пропускаем
            if pval is not None:
                c['edgePerM16' if plate == '16' else 'edgePerM32'] = pval
        elif tag == 'dfill':
            fills = data['slidingDoor']['fills']
            if parts[1] == 'mirror':
                if pval is not None:
                    fills['mirror']['pricePerM2'] = pval
                if edited(extra, base, 'color'):     # название зеркала — в колонке «Цвет»
                    fills['mirror']['name'] = edited(extra, base, 'color')
            elif parts[1] in ('glass', 'extra'):
                if parts[1] == 'glass':
                    typ, hit = fills['glass'], next((g for g in fills['glass']['colors']
                                                     if g['id'] == parts[2]), None)
                else:                                 # заведённый пользователем тип наполнения
                    typ = next((z for z in fills.get('extra', []) if z['id'] == parts[2]), None)
                    hit = next((c for c in typ['colors'] if c['id'] == parts[3]), None) if typ else None
                if hit is None:
                    return f'{errctx}: не найдена позиция наполнения {key}'
                if pval is not None:
                    hit['pricePerM2'] = pval
                if edited(extra, base, 'color'):
                    hit['name'] = edited(extra, base, 'color')
                if edited(extra, base, 'hex'):
                    hit['color'] = edited(extra, base, 'hex')
                if edited(extra, base, 'name'):       # имя ТИПА повторяется в каждой его строке
                    typ['name'] = edited(extra, base, 'name')
        elif tag == 'prof':
            _, el, colr = parts
            hit = next((p for p in data['slidingDoor']['profilePrices']
                        if p['element'] == el and p['color'] == colr), None)
            if hit is None:
                return f'{errctx}: не найден профиль {el}×{colr}'
            if pval is not None:
                hit['pricePerM'] = pval
            # «Цвет» и «Цвет (hex)» — это цвет профиля (лист «Цвета профилей» убран, правится здесь);
            # цвет общий для всех элементов, поэтому смотрим только РЕАЛЬНО изменённые ячейки.
            col = next((c for c in data['slidingDoor']['colors'] if c['id'] == colr), None)
            if col is not None:
                if edited(extra, base, 'color'):
                    col['name'] = edited(extra, base, 'color')
                if edited(extra, base, 'hex'):
                    col['hex'] = edited(extra, base, 'hex')
            nm = edited(extra, base, 'name')          # «Название» = элемент профиля (тоже общее)
            if nm:
                if el in ELEMENT_LABELS:
                    set_label(data, f'profel:{el}', nm, ELEMENT_LABELS[el])
                else:                                 # вертикальный — это имя профиля из каталога
                    p = next((x for x in data['slidingDoor']['profiles'] if x['id'] == el), None)
                    base = re.sub(r'\s*вертикальный$', '', nm, flags=re.IGNORECASE).strip()
                    if p is not None and base:
                        p['name'] = base
        elif tag == 'profcol':
            cid = parts[1]
            hit = next((c for c in data['slidingDoor']['colors'] if c['id'] == cid), None)
            if hit is None:
                return f'{errctx}: не найден цвет профиля {cid}'
            if extra.get('color'):        # название цвета профиля — в колонке «Цвет»
                hit['name'] = str(extra['color'])
            if extra.get('hex'):
                hit['hex'] = str(extra['hex'])
        elif tag == 'mesh':
            # ключ = mesh:<gid>; старый вид mesh:<глубина>:<цвет> = тот же gid через «_»
            gid = '_'.join(parts[1:])
            hit = by_gid(data['meshShelf'], gid)
            if hit is None:
                return f'{errctx}: не найдена сетчатая полка {key}'
            if pval is not None:
                hit['pricePerM'] = pval
            if txt(extra.get('name')):
                hit['name'] = txt(extra.get('name'))
            dep = num(extra.get('w'))                  # глубина полки — колонка «Ширина, мм»
            if dep is not None:
                hit['depth'] = int(dep)
            colr = REV_METAL.get(str(extra.get('color') or '').strip().lower())
            if colr:
                hit['color'] = colr
            if clash(data['meshShelf'], hit, ('depth', 'color')):
                return f'{errctx}: такая сетчатая полка уже есть (глубина + цвет)'
        elif tag == 'basket':
            gid = '_'.join(parts[1:])
            hit = by_gid(data['basket'], gid)
            if hit is None:
                return f'{errctx}: не найдена корзина {key}'
            if pval is not None:
                hit['price'] = pval
            for fld, col in (('height', 'h'), ('width', 'l'), ('depth', 'w')):
                v = num(extra.get(col))
                if v is not None:
                    hit[fld] = int(v)
            colr = REV_METAL.get(str(extra.get('color') or '').strip().lower())
            if colr:
                hit['color'] = colr
            if clash(data['basket'], hit, ('width', 'depth', 'height', 'color')):
                return f'{errctx}: такая корзина уже есть (размеры + цвет)'
            set_label(data, f'basket:{gid}', extra.get('name'), 'Корзина')
        elif tag == 'slide':
            gid = '_'.join(parts[1:])
            hit = by_gid(data['drawerSlide'], gid)
            if hit is None:
                return f'{errctx}: не найдена направляющая {key}'
            if pval is not None:
                hit['price'] = pval
            ln = num(extra.get('l'))
            if ln is not None:
                hit['length'] = int(ln)
            if clash(data['drawerSlide'], hit, ('type', 'length')):
                return f'{errctx}: такая направляющая уже есть (тип + длина)'
            set_label(data, f'slide:{gid}', extra.get('name'),
                      SLIDE_TYPES.get(hit['type'], hit['type']) + ' направляющие')
        # ── Не добавляемые позиции (верхние таблицы «Фурнитуры» и «Услуг»): только цена ──────────
        elif tag == 'fit':
            fid = parts[1]
            hit = next((it for it in data['fittings'] if it['id'] == fid), None)
            if hit is None:
                return f'{errctx}: не найдена фурнитура {fid}'
            if pval is not None:
                hit['price'] = pval
        elif tag in ('swing', 'rollers', 'rod', 'softclose', 'handle'):
            # одиночные позиции фурнитуры: у каждой своё поле цены
            obj, fld = {'swing': (data['swingDoorHardware'], 'pricePerDoor'),
                        'rollers': (data['slidingDoor']['rollers'], 'pricePerSet'),
                        'rod': (data.get('rod'), 'pricePerM'),
                        'softclose': (data.get('doorSoftClose'), 'pricePerDoor'),
                        'handle': (data.get('drawerHandle'), 'pricePerDrawer')}[tag]
            if obj is None:
                return f'{errctx}: не найдена позиция {key}'
            if pval is not None:
                obj[fld] = pval
        elif tag == 'service':
            sv = data.setdefault('services', {}).setdefault(parts[1], {})
            if pval is not None:
                sv['price'] = pval
        # ── Добавляемые позиции (нижние таблицы): раздел — в «От чего зависит» ───────────────────
        elif tag == 'fitopt':
            hit = next((i for i in data.get('fittingOptions', []) if i['gid'] == parts[1]), None)
            if hit is None:
                return f'{errctx}: не найдена позиция фурнитуры {key}'
            if pval is not None:
                hit['price'] = pval
            if txt(extra.get('name')):
                hit['name'] = txt(extra.get('name'))
            if txt(extra.get('dep')):                 # правка раздела = перенос позиции в него
                hit['group'] = txt(extra.get('dep'))
            if txt(extra.get('unit')):
                unit = fit_unit(extra)
                if unit is None:
                    return f'{errctx}: ед.изм «{txt(extra.get("unit"))}» — допустимы только ' + ', '.join(UNITS_FIT)
                hit['unit'] = unit
        elif tag == 'addon':
            _, grp, item = parts
            g = next((x for x in data['extras'] if x['id'] == grp), None)
            it = next((y for y in g['items'] if y['id'] == item), None) if g else None
            if it is None:
                return f'{errctx}: не найден доп.элемент {key}'
            if manual:                                # «вручную» ↔ число: переключаем сам режим
                it['manual'] = True
            else:
                it.pop('manual', None)
                it['price'] = pval
            if txt(extra.get('name')):                # позиция — в «Названии»
                it['name'] = txt(extra.get('name'))
            if edited(extra, base, 'dep'):            # раздел повторяется в каждой строке группы
                g['name'] = edited(extra, base, 'dep')
        else:
            return f'{errctx}: неизвестный тип ключа «{tag}»'
    except (KeyError, ValueError, IndexError) as e:
        return f'{errctx}: сбой применения {key} ({e})'
    return None


def main():
    global IN_DIR
    import openpyxl
    args = sys.argv[1:]
    path = None
    if args and args[0] == '--in':
        path = args[1]
    else:
        if len(args) >= 2 and args[0] == '--dir':  # папка, где искать файл (батник: Config\Выгрузки)
            IN_DIR = args[1]
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
            path = filedialog.askopenfilename(title='Выберите файл выгрузки для загрузки',
                                              initialdir=IN_DIR, filetypes=[('Excel', '*.xlsx')])
            root.destroy()
        except Exception:
            pass
    if not path or not os.path.exists(path):
        print('Загрузка отменена — файл не выбран.'); return 1

    with open(DST, encoding='utf-8') as f:
        original = json.load(f)
    data = copy.deepcopy(original)
    ensure_gids(data)  # проставить стабильные gid позициям (для правки по ключу)
    baseline = build_baseline(original)  # как строки выглядели в выгрузке — чтобы видеть, что правили

    wb = openpyxl.load_workbook(path, data_only=True)
    errors, applied, created, skipped_new = [], 0, 0, 0

    for name in SHEET_NAMES:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v in (None, '') for v in row):
                continue
            # Единый формат: колонки одинаковы на всех листах (см. COLS).
            extra = {f: (row[c - 1] if len(row) >= c else None) for f, c in COLS.items()}
            key = extra['key']
            if str(key or '').startswith('#'):
                continue  # серая строка-заголовок таблицы внутри листа
            if not key:
                # новая строка (без ключа) — создать позицию, если лист это допускает
                creator = CREATORS.get(name)
                if not creator:
                    skipped_new += 1
                    continue
                err = creator(data, extra, f'{name}, строка {ri} (новая)')
                if err:
                    errors.append(err)
                else:
                    created += 1
                continue
            err = apply_row(data, str(key), extra['price'], extra, f'{name}, строка {ri}',
                            baseline.get(str(key)))
            if err:
                errors.append(err)
            else:
                applied += 1
                store_meta(data, str(key), extra, baseline.get(str(key)))

    ensure_gids(data)  # позиции, добавленные новыми строками, тоже получают стабильный id
    if not data.get('catalogMeta'):  # не оставлять пустую секцию в файле
        data.pop('catalogMeta', None)

    if errors:
        print(f'ЗАГРУЗКА ОТМЕНЕНА — {len(errors)} ошибок, файл НЕ изменён:')
        for e in errors[:50]:
            print('  • ' + e)
        return 1

    # backup + атомарная запись
    try:
        with open(DST + '.bak', 'w', encoding='utf-8') as f:
            json.dump(original, f, ensure_ascii=False, indent=2)
        tmp = DST + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, DST)
    except Exception as e:
        print(f'ОШИБКА записи: {e}'); return 1

    msg = f'Готово: обновлено {applied}, добавлено новых {created}'
    if skipped_new:
        msg += f', пропущено новых строк на нерасширяемых листах {skipped_new}'
    print(msg + '.')
    print(f'Бэкап прежней версии: {os.path.basename(DST)}.bak')
    return 0


if __name__ == '__main__':
    code = main()
    if not (len(sys.argv) > 1 and sys.argv[1] == '--in'):
        input('\nНажмите Enter, чтобы закрыть...')
    sys.exit(code)
