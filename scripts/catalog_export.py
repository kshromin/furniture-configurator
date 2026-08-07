# -*- coding: utf-8 -*-
# Выгрузка ассортимента/цен конфигуратора в Excel — ЕДИНЫЙ формат колонок (шаблон пользователя
# «Выгрузки/шаблон.xlsx», задание 4.08). Категории = отдельные листы (вариант «А»), но набор и
# порядок колонок ОДИНАКОВЫЙ на всех листах — что не относится к категории, остаётся пустым.
#
# Колонки (см. HEADERS): ключ | производитель | название | цвет | ед.изм | ЦЕНА | высота | длинна |
# ширина | от чего зависит | цвет(hex) | Корпус | Фасад | Наполнение | файл текстуры.
# Ключ теперь ВИДИМЫЙ (первая колонка) — по нему загрузка находит позицию.
# Ручная цена = слово «вручную». Данные — из data/materials.json.
#
# Поля, которых в базе нет (производитель у неплитных категорий, габариты листа/хлыста и т.п.),
# хранятся в data['catalogMeta'][ключ] = {producer,h,l,w,dep,hex}: в первой выгрузке они пустые,
# пользователь заполняет в Excel, загрузка сохраняет обратно. Поля, которые СЛЕДУЮТ из самой
# позиции (толщина плиты, длина направляющей, размеры корзины, глубина сетки, hex профиля…),
# выводятся из базы и в meta не пишутся — см. DERIVED и «Справку» в книге; правятся они прямо в
# базе (ключ тот же = та же позиция).
#
# КЛЮЧ = СТАБИЛЬНЫЙ ID позиции (задание «формат выгрузки»): пока ключ в строке тот же, правка ЛЮБОГО
# столбца обновляет ту же позицию (дубля не будет); строка без ключа = новая позиция. Поэтому у
# категорий, где ключ раньше собирался из атрибутов (сетки/корзины/направляющие, ЛДСП), он теперь
# строится из `gid` — см. gid_of(). Названия позиций, которых нет в базе (направляющие, общие
# элементы профиля), переопределяются через data['catalogLabels'] — см. label_of().
#
# Запуск:
#   python catalog_export.py                 → окно с галочками (какие категории) + окно «куда сохранить»
#   python catalog_export.py --all <path>    → без окон: все категории в <path> (для тестов/автоматизации)
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'data', 'materials.json')
OUT_DIR = os.path.normpath(os.path.join(ROOT, '..', 'Выгрузки'))
OUT = os.path.join(OUT_DIR, 'ассортимент.xlsx')

SURFACES = [('korpus', 'корпус'), ('fasad', 'фасад'), ('fill', 'наполнение')]
SLIDE_TYPES = {'ball': 'Шариковые', 'soft': 'С доводчиком', 'push': 'Push-to-open', 'blum': 'BLUM'}
METAL_COLORS = {'white': 'Белый', 'silver': 'Серебро', 'black': 'Чёрный'}
MANUAL = 'вручную'  # маркер ручной цены (DATA-SCHEMA: price = "manual")

# Общие для всех листов колонки (порядок менять нельзя — по нему читает catalog_import.py).
HEADERS = ['Удалить', 'Артикул/ид/key', 'Производитель', 'Название', 'Цвет', 'Ед.изм', 'Цена',
           'Высота, мм', 'Длинна, мм', 'Ширина, мм', 'От чего зависит', 'Цвет (hex)',
           'Корпус', 'Фасад', 'Наполнение', 'Файл текстуры']
WIDTHS = [9, 10, 16, 24, 24, 10, 10, 12, 12, 12, 30, 12, 8, 8, 12, 18]  # «Удалить» и ключ узкие
PRICE_COL = 7      # сдвинулась на 1: слева добавлен столбец «Удалить» (7.08)

# Единицы измерения
U_M2, U_M, U_PC, U_SET, U_PART = 'кв.м', 'пог.м', 'шт', 'комплект', 'деталь'
PER_UNIT = {'item': 'изделие', 'shelf': 'полка', 'rod': 'штанга', 'rodFlange': 'фланец',
            'set': U_SET, 'door': 'дверь', 'drawer': 'ящик'}
# Названия «общих» элементов профиля (вертикальные берут имя из каталога профилей)
ELEMENT_LABELS = {'horizTop': 'Горизонт верхний', 'horizBottom': 'Горизонт нижний',
                  'divider': 'Перемычка', 'track': 'Направляющая (верх+низ)'}

# Какие поля СЛЕДУЮТ из самой позиции (не из catalogMeta): их правка в Excel не сохраняется —
# позиция задаётся ключом. Ключ таблицы — тег ключа (часть до первого «:»).
DERIVED = {
    'ldsp':   {'h', 'hex', 'dep', 'producer', 'l', 'w', 'color', 'texture'},  # всё пишется в сам материал: толщина/hex/производитель, макс. размер детали, имя цвета, текстура
    'ldspm':  {'h', 'hex', 'producer', 'color', 'texture'},  # высота = толщина плиты, hex/производитель — из базы
    'edge':   {'h', 'dep', 'color'},  # высота = плита 16/32, «от чего зависит» = ключ ЛДСП, цвет = имя материала
    'dfill':  {'hex', 'color'},
    'prof':   {'hex', 'color'},      # hex и название цвета — каталог цветов профиля
    'profcol': {'hex', 'color'},
    'hdf':    {'h', 'hex', 'color', 'dep', 'producer', 'texture'},  # всё в самой позиции; «зависит» — подпись
    'addon':  {'dep', 'producer'},   # раздел = группа extras, производитель — поле самой позиции
    'fitopt': {'dep', 'unit', 'producer'},  # назначение, ед.изм и бренд хранятся в самой позиции
    'mesh':   {'w', 'color'},        # ширина = глубина полки, цвет = цвет полки
    'basket': {'h', 'l', 'w', 'color'},  # высота/длина/ширина и цвет — свойства корзины
    'slide':  {'l'},                 # длина направляющей
}

# Листы «Фурнитура»/«Услуги»/«Доп.элементы» делятся на ДВЕ таблицы (задание 5.08): сверху позиции,
# заведённые в самой программе (их нельзя добавлять/удалять — можно менять только цену), снизу
# добавляемые, где раздел позиции задаётся колонкой «От чего зависит».
BAND_FIXED = 'НЕ ДОБАВЛЯЕМЫЕ — заведены в программе; менять можно ТОЛЬКО цену'
BAND_ADD_FIT = ('ДОБАВЛЯЕМЫЕ — «От чего зависит» = НАЗНАЧЕНИЕ, только из списка в ячейке (пусто или '
                'своё слово = строка не загрузится). «Название» = что это, «Ед.изм» = шт/комплект/пог.м')
# Назначения добавляемой фурнитуры — ровно то, что конфигуратор УМЕЕТ выбирать. Список закрытый:
# пользователь выбирает его из выпадашки (орфография исключена), а загрузка отвергает всё
# остальное — иначе в каталоге копится фурнитура, которую негде выбрать (просьба 6.08).
# ⚠️ Появится новое назначение в интерфейсе — добавить сюда И в правило поиска (см. HANDLE_GROUP_RE
# в js/core/materials.js: раздел ищется по слову «ручк»).
FIT_PURPOSES = ('Ручки ящика',)
BAND_ADD_SERV = ('ДОБАВЛЯЕМЫЕ — новая строка снизу заводит услугу: «Название» = что это, '
                 '«От чего зависит» = раздел (Доставка, Монтаж, Подъём на этаж…)')
BAND_ADD_EXTRA = ('ДОБАВЛЯЕМЫЕ — новая строка снизу заводит позицию: «Название» = что это, '
                  '«От чего зависит» = раздел')
BAND_ADD_FILL = ('«Название» = ТИП наполнения (Зеркало, Стекло, Ротанг…), «Цвет» = позиция этого '
                 'типа. Новая строка снизу: знакомый тип — добавит позицию, новый — заведёт тип')
# Группы `extras`, которые по смыслу услуги (уходят на лист «Услуги»), пока в данных не проставлен
# явный `kind`. Первая загрузка проставит `kind` каждой группе, дальше берётся оттуда.
DEFAULT_SERVICE_GROUPS = {'delivery', 'lift', 'montage'}

HELP_TEXT = [
    'Ассортимент и цены — как пользоваться (единый формат, 4.08)',
    '',
    '1. Колонки ОДИНАКОВЫЕ на всех листах; что не относится к категории — пустое. Порядок не менять.',
    '2. Каждая строка — одна позиция. «Артикул/ид/key» — по нему загрузка находит позицию, НЕ МЕНЯТЬ.',
    '3. Цена — ЖЁЛТАЯ колонка (числом, без «₽» и пробелов). «вручную» = цену вводит пользователь при заказе.',
    '4. Ключ тот же — значит ТА ЖЕ позиция: правьте ЛЮБОЙ столбец (название, цвет, размер, цена),',
    '   загрузка обновит эту же позицию, дубля не будет. Ключ стёрт/пустой = НОВАЯ позиция.',
    '5. Пустые «Производитель», габариты (высота/длинна/ширина), «От чего зависит» — заполняйте:',
    '   они сохранятся при загрузке (в базе таких полей раньше не было).',
    '6. Новая строка внизу листа без ключа = новая позиция (ЛДСП, наполнение дверей, сетки, корзины,',
    '   доп.элементы). На остальных листах новые строки пропускаются.',
    '7. ЛДСП: одна строка = один материал; «Корпус/Фасад/Наполнение» — да/нет, где он доступен.',
    '   «да» без записи — заведётся, «нет» с записью — уберётся. Вид задаёт «Файл текстуры».',
    '   «Высота» = толщина плиты; «Длинна» и «Ширина» = МАКСИМАЛЬНО допустимый размер детали из',
    '   этого материала (можно оставить пустыми, если ограничения нет).',
    '8. Кромка привязана к материалу ЛДСП — см. «От чего зависит» (ключ ЛДСП) и высоту (плита 16/32).',
    '9. Наполнение дверей: «Название» = ТИП (Зеркало, Стекло, Ротанг…), «Цвет» = позиция типа.',
    '   Новый тип можно завести строкой снизу; конфигуратор пока предлагает ЛДСП, зеркало, стекло',
    '   и спеццвет — новые типы копятся в каталоге, пока не сделаем их выбор в интерфейсе.',
    '10. Листы «Фурнитура», «Услуги», «Доп.элементы» — ДВЕ таблицы под серыми заголовками:',
    '   • НЕ ДОБАВЛЯЕМЫЕ (сверху) — заведены в самой программе; правится ТОЛЬКО цена,',
    '     остальные колонки в этой таблице загрузка не читает;',
    '   • ДОБАВЛЯЕМЫЕ (снизу) — ваш ассортимент: «Название» = что это, «От чего зависит» = раздел.',
    '     Новая строка с уже знакомым разделом добавит позицию в него, с новым — заведёт раздел.',
    '   Серые строки-заголовки (начинаются с «#») не трогайте — загрузка их пропускает.',
    '11. ЦВЕТ (hex) руками задавать не нужно: он считается САМ из картинки декора, когда вы',
    '    заливаете текстуры («Загрузить текстуры.bat»). Задавайте его только для однотонных',
    '    материалов без картинки — впишите код вида #a1b2c3 ИЛИ просто закрасьте ячейку.',
    '12. УДАЛИТЬ позицию: в КРАЙНЕМ ЛЕВОМ столбце «Удалить» выберите «да» (строку не стирайте —',
    '    чего нет в файле, то загрузка просто не трогает). Удалять можно только СВОИ позиции: то,',
    '    что заведено в самой программе (верхние таблицы), не удаляется. Отчёт покажет, что удалено.',
    '13. Загружать можно ЧАСТЬ книги: лишние строки и листы можно удалить из файла — позиции,',
    '    которых в нём нет, останутся в базе нетронутыми.',
    '14. Когда закончили — сохраните файл и запустите загрузку («Загрузить цены.bat»).',
]


def load():
    with open(SRC, encoding='utf-8') as f:
        return json.load(f)


def meta_of(d, key):
    return (d.get('catalogMeta') or {}).get(key, {})


def row(d, key, name='', color='', unit='', price='', producer='', h='', l='', w='',
        dep='', hexv='', korpus='', fasad='', fill='', texture=''):
    """Строка в едином формате. Пустые «свободные» поля добираются из catalogMeta по ключу —
    так вписанное в Excel не пропадает, даже если приложение это поле пока не использует."""
    m = meta_of(d, key)
    return ['',                       # «Удалить» — пусто; «да» в этой ячейке = убрать позицию
            key,
            producer or m.get('producer', ''),
            name,
            color or m.get('color', ''),
            m.get('unit') or unit,
            price,
            h if h != '' else m.get('h', ''),
            l if l != '' else m.get('l', ''),
            w if w != '' else m.get('w', ''),
            dep or m.get('dep', ''),
            hexv or m.get('hex', ''),
            korpus, fasad, fill, texture or m.get('texture', '')]


# ── Построители категорий: возвращают dict(title, rows) ─────────────────────────────────────

def ldsp_gid_of(prod, c):
    """Стабильный id материала (НЕ из имени): свой gid, иначе детерминированно из id первого члена."""
    return c.get('gid') or f"{prod['id']}__{c['id']}"


def gid_of(item, *fields):
    """Стабильный id позиции: свой gid, иначе детерминированно из полей (совпадает со СТАРЫМ ключом,
    поэтому прежние выгрузки грузятся без потерь). Нужен, чтобы правка любого столбца — размера,
    цвета, названия — обновляла ТУ ЖЕ позицию, а не заводила новую (задание «формат выгрузки»)."""
    return item.get('gid') or '_'.join(str(item[f]) for f in fields)


def label_of(d, key, default):
    """Название позиции, у которой нет поля имени в базе (направляющие, общие элементы профиля):
    правка из Excel хранится в data['catalogLabels'][ключ]."""
    return (d.get('catalogLabels') or {}).get(key) or default


def ldsp_kind_disp(col):
    """Тип материала и его название по отдельным колонкам («Название» и «Цвет»).
    Тип хранится в самом материале (`kind`); у старых записей, где поля ещё нет, выводим его из
    имени («ЛДСП Белый альпийский» → тип «ЛДСП», цвет «Белый альпийский»)."""
    name = str(col.get('name') or '')
    kind = str(col.get('kind') or '').strip()
    if not kind and name.lower().startswith('лдсп '):
        kind = 'ЛДСП'
    if kind and name.lower().startswith(kind.lower()):
        name = name[len(kind):].strip()
    return kind, name


def cat_ldsp(d):
    # Одна строка на материал (группа по производитель+имя+толщина через все поверхности);
    # поверхности — столбцы да/нет. Ключ = стабильный gid (не из имени) → правка имени = обновление
    # той же позиции, без дублей. «Название» = ЛДСП, «Цвет» = название цвета без слова «ЛДСП».
    groups = {}  # (произв., имя, толщина) -> {gid, price, texture, hex, cname, surf:set(), order}
    seq = 0
    for surf, _lbl in SURFACES:
        for prod in d[surf]['producers']:
            for c in prod['colors']:
                k = (prod['name'], c['name'], c.get('thickness', 16))
                g = groups.get(k)
                if g is None:
                    g = {'gid': ldsp_gid_of(prod, c), 'price': c['pricePerM2'],
                         'texture': c.get('texture', ''), 'hex': c.get('color', ''),
                         'cname': c['name'], 'kind': c.get('kind', ''), 'surf': set(), 'order': seq,
                         # макс. допустимый размер детали из этого материала (задание 5.08)
                         'maxl': c.get('maxPartL', ''), 'maxw': c.get('maxPartW', '')}
                    seq += 1
                    groups[k] = g
                else:
                    if c.get('gid'):
                        g['gid'] = c['gid']  # сохранённый gid приоритетнее выведенного
                    if not g['kind'] and c.get('kind'):
                        g['kind'] = c['kind']
                    for fld, src in (('maxl', 'maxPartL'), ('maxw', 'maxPartW')):
                        if g[fld] == '' and c.get(src) not in (None, ''):
                            g[fld] = c[src]
                g['surf'].add(surf)
    yn = lambda s, g: 'да' if s in g['surf'] else 'нет'
    rows = []
    for (pname, cname, th), g in sorted(groups.items(), key=lambda kv: (str(kv[0][0]).lower(), str(kv[0][1]).lower(), kv[0][2])):
        key = f"ldsp:{g['gid']}"
        # «Название» = «ЛДСП» ТОЛЬКО если материал реально ЛДСП; иначе пусто, а «Цвет» = полное имя
        # (не-ЛДСП, напр. «Этерно МДФ Белый», не трогаем — иначе при загрузке приклеится «ЛДСП»).
        kind, disp = ldsp_kind_disp({'name': cname, 'kind': g['kind']})
        rows.append(row(d, key, name=kind, color=disp,
                        unit=U_M2, price=g['price'], producer=pname, h=th,
                        l=g['maxl'], w=g['maxw'], hexv=g['hex'],
                        korpus=yn('korpus', g), fasad=yn('fasad', g), fill=yn('fill', g), texture=g['texture']))
    # ХДФ — материал ТОЛЬКО задней стенки (в корпус/фасад/наполнение не идёт, поэтому да/нет
    # у него пустые). Одна позиция, правится цена/толщина/цвет — см. задание «разобр с хдф».
    hdf = d.get('hdf') or {}
    for c in hdf.get('colors', []):
        rows.append(row(d, f"hdf:{c['id']}", name=hdf.get('kind', 'ХДФ'), color=c.get('name', ''),
                        unit=U_M2, price=c.get('pricePerM2', 0), h=c.get('thickness', 4),
                        producer=c.get('producer', ''), hexv=c.get('color', ''),
                        dep='Задняя стенка', texture=c.get('texture', '')))
    return dict(title='МАТЕРИАЛ', rows=rows)


def cat_edge(d):
    # Цвет — без слова «ЛДСП» (как у ЛДСП). Сортировка ПО ВЫСОТЕ (все 16, потом все 32).
    tmp = []  # (plate, row)
    for surf, _label in SURFACES:
        for prod in d[surf]['producers']:
            for c in prod['colors']:
                for plate, field in ((16, 'edgePerM16'), (32, 'edgePerM32')):
                    if field in c:
                        key = f"edge:{surf}:{prod['id']}:{c['id']}:{plate}"
                        tmp.append((plate, row(d, key, name='Кромка', color=ldsp_kind_disp(c)[1],
                                    unit=U_M, price=c[field], h=plate,
                                    dep=f"ldsp:{ldsp_gid_of(prod, c)}")))
    rows = [r for _plate, r in sorted(tmp, key=lambda x: x[0])]
    return dict(title='Кромка', rows=rows)


def cat_door_fill(d):
    # «Название» = ТИП наполнения (Зеркало, Стекло, Ротанг…), «Цвет» = позиция этого типа.
    # Новая строка с новым типом заводит тип (хранится в fills['extra'], см. DATA-SCHEMA).
    fills = d['slidingDoor']['fills']
    rows = [band(BAND_ADD_FILL),
            row(d, 'dfill:mirror', name='Зеркало', color=fills['mirror'].get('name', 'Зеркало'),
                unit=U_M2, price=fills['mirror']['pricePerM2'])]
    glass_type = fills.get('glass', {}).get('name', 'Стекло')
    for c in fills.get('glass', {}).get('colors', []):
        rows.append(row(d, f"dfill:glass:{c['id']}", name=glass_type, color=c['name'], unit=U_M2,
                        price=c['pricePerM2'], hexv=c.get('color', '')))
    for t in fills.get('extra', []):
        for c in t.get('colors', []):
            rows.append(row(d, f"dfill:extra:{t['id']}:{c['id']}", name=t['name'], color=c['name'],
                            unit=U_M2, price=c.get('pricePerM2', 0), hexv=c.get('color', '')))
    # шаблон ручной позиции «спеццвет» (единая механика manual)
    rows.append(row(d, 'dfill:special', name=glass_type, color='Спеццвет (ручная цена)',
                    unit=U_M2, price=MANUAL))
    return dict(title='Наполнение дверей', rows=rows)


def cat_profile(d):
    prof = {p['id']: p['name'] for p in d['slidingDoor']['profiles']}
    cols = {c['id']: c for c in d['slidingDoor']['colors']}
    rows = []
    for pp in d['slidingDoor'].get('profilePrices', []):
        el = pp['element']
        # Вертикальные — имя профиля из каталога; общие (горизонты/перемычка/направляющая) — метка,
        # которую можно переименовать из Excel (хранится в catalogLabels['profel:<элемент>']).
        label = (label_of(d, f'profel:{el}', ELEMENT_LABELS[el]) if el in ELEMENT_LABELS
                 else prof.get(el, el) + ' вертикальный')
        c = cols.get(pp['color'], {})
        rows.append(row(d, f"prof:{el}:{pp['color']}", name=label, color=c.get('name', pp['color']),
                        unit=U_M, price=pp['pricePerM'], hexv=c.get('hex', '')))
    return dict(title='Профили купе', rows=rows)


def cat_profile_colors(d):
    rows = [row(d, f"profcol:{c['id']}", name='Цвет профиля', color=c['name'], hexv=c.get('hex', ''))
            for c in d['slidingDoor']['colors']]
    return dict(title='Цвета профилей', rows=rows)


def cat_mesh(d):
    rows = []
    for m in d['meshShelf']:
        key = f"mesh:{gid_of(m, 'depth', 'color')}"
        rows.append(row(d, key, name=m['name'], color=METAL_COLORS.get(m['color'], m['color']),
                        unit=U_M, price=m['pricePerM'], w=m['depth']))
    return dict(title='Сетчатые полки', rows=rows)


def cat_basket(d):
    rows = []
    for b in d['basket']:
        key = f"basket:{gid_of(b, 'width', 'depth', 'height', 'color')}"
        rows.append(row(d, key, name=label_of(d, key, 'Корзина'),
                        color=METAL_COLORS.get(b['color'], b['color']), unit=U_PC, price=b['price'],
                        h=b['height'], l=b['width'], w=b['depth']))
    return dict(title='Корзины', rows=rows)


def cat_slide(d):
    rows = []
    for s in d['drawerSlide']:
        key = f"slide:{gid_of(s, 'type', 'length')}"
        default = SLIDE_TYPES.get(s['type'], s['type']) + ' направляющие'
        rows.append(row(d, key, name=label_of(d, key, default), unit=U_PC, price=s['price'],
                        l=s['length']))
    return dict(title='Направляющие', rows=rows)


def band(text):
    """Строка-заголовок таблицы внутри листа (склеивается на всю ширину; загрузка её пропускает)."""
    return ['# ' + text] + [''] * (len(HEADERS) - 1)


BLANK = [''] * len(HEADERS)


def group_kind(grp):
    """Группа `extras` — услуга («Услуги») или доп.элемент («Доп.элементы»)."""
    return grp.get('kind') or ('service' if grp['id'] in DEFAULT_SERVICE_GROUPS else 'extra')


def addon_rows(d, kind):
    """Добавляемые позиции из `extras` нужного вида: «Название» = позиция, раздел = группа."""
    rows = []
    for grp in d['extras']:
        if group_kind(grp) != kind:
            continue
        for it in grp['items']:
            key = f"addon:{grp['id']}:{it['id']}"
            price = MANUAL if it.get('manual') else it['price']
            rows.append(row(d, key, name=it['name'], unit=U_PC, price=price, dep=grp['name'],
                            producer=it.get('producer', '')))
    return rows


def cat_hardware(d):
    # Ед.изм — только шт / комплект / пог.м (других нет). Штука по умолчанию.
    rows = [band(BAND_FIXED)]
    for it in d['fittings']:
        rows.append(row(d, f"fit:{it['id']}", name=it['name'], unit=U_PC, price=it['price']))
    sw = d['swingDoorHardware']
    rows.append(row(d, 'swing', name=sw['name'], unit=U_SET, price=sw['pricePerDoor']))
    ro = d['slidingDoor']['rollers']
    rows.append(row(d, 'rollers', name=ro['name'], unit=U_SET, price=ro['pricePerSet']))
    rod = d.get('rod')
    if rod:
        rows.append(row(d, 'rod', name=rod['name'], unit=U_M, price=rod['pricePerM']))
    sc = d.get('doorSoftClose')
    if sc:
        rows.append(row(d, 'softclose', name=sc['name'], unit=U_PC, price=sc['pricePerDoor']))
    dh = d.get('drawerHandle')
    if dh:
        rows.append(row(d, 'handle', name=dh['name'], unit=U_PC, price=dh['pricePerDrawer']))
    # Вторая таблица: фурнитура, которую ведёт пользователь. Назначение («От чего зависит») —
    # выбором из FIT_PURPOSES; чтобы значение было под рукой, каждое назначение печатается
    # строкой-образцом (без ключа и цены — она ничего не заводит, из неё копируют ячейку вниз).
    rows += [BLANK, band(BAND_ADD_FIT)]
    opts = d.get('fittingOptions', [])
    for purpose in FIT_PURPOSES:
        rows.append(row(d, '', name='', unit='', price='', dep=purpose))
        for it in opts:
            if str(it.get('group', '')).strip().lower() == purpose.lower():
                rows.append(row(d, f"fitopt:{it['gid']}", name=it['name'], unit=it.get('unit', U_PC),
                                price=it['price'], dep=it.get('group', ''),
                                producer=it.get('producer', '')))
    for it in opts:   # назначения, которых больше нет в списке (остались с прежних выгрузок)
        if str(it.get('group', '')).strip().lower() not in {p.lower() for p in FIT_PURPOSES}:
            rows.append(row(d, f"fitopt:{it['gid']}", name=it['name'], unit=it.get('unit', U_PC),
                            price=it['price'], dep=it.get('group', ''),
                            producer=it.get('producer', '')))
    return dict(title='Фурнитура', rows=rows)


def cat_service(d):
    # Услуги: сверху заведённые в программе (считаются по её правилам — Крепёж, Встройка),
    # снизу добавляемые (Доставка, Монтаж, Подъём… — группы `extras` вида «услуга»).
    rows = [band(BAND_FIXED)]
    for sid, sv in (d.get('services') or {}).items():
        rows.append(row(d, f'service:{sid}', name=sv.get('name', sid), unit=U_PART, price=sv.get('price', 0)))
    rows += [BLANK, band(BAND_ADD_SERV)] + addon_rows(d, 'service')
    return dict(title='Услуги', rows=rows)


def cat_addon(d):
    # Доп.элементы — все добавляемые (не добавляемых тут нет).
    rows = [band(BAND_ADD_EXTRA)] + addon_rows(d, 'extra')
    # шаблон ручной «заказной» позиции
    rows.append(row(d, 'addon:custom:manual', name='Заказная позиция (ручная цена)',
                    unit=U_PC, price=MANUAL, dep='Доп. элементы'))
    return dict(title='Доп.элементы', rows=rows)


# Порядок = порядок в диалоге и в книге. (key, человекочитаемое имя, builder)
CATEGORIES = [
    ('ldsp', 'МАТЕРИАЛ (корпус/фасад/наполнение)', cat_ldsp),
    ('edge', 'Кромка', cat_edge),
    ('door_fill', 'Наполнение дверей', cat_door_fill),
    ('profile', 'Профили купе', cat_profile),
    ('mesh', 'Сетчатые полки', cat_mesh),
    ('basket', 'Корзины', cat_basket),
    ('slide', 'Направляющие', cat_slide),
    ('hardware', 'Фурнитура', cat_hardware),
    ('service', 'Услуги', cat_service),
    ('addon', 'Доп.элементы', cat_addon),
]


def choose_categories():
    """Окно с галочками: какие категории выгрузить. Возвращает список ключей или None (отмена)."""
    try:
        import tkinter as tk
    except Exception:
        return [k for k, _, _ in CATEGORIES]
    root = tk.Tk()
    root.title('Что выгрузить')
    root.attributes('-topmost', True)
    tk.Label(root, text='Отметьте категории для выгрузки:', font=('Segoe UI', 10, 'bold')).pack(anchor='w', padx=12, pady=(12, 4))
    vars_ = {}
    for k, label, _ in CATEGORIES:
        v = tk.BooleanVar(value=True)
        vars_[k] = v
        tk.Checkbutton(root, text=label, variable=v).pack(anchor='w', padx=18)
    result = {'ok': False}
    def ok():
        result['ok'] = True; root.quit()
    tk.Button(root, text='Выгрузить', width=16, command=ok).pack(pady=12)
    root.protocol('WM_DELETE_WINDOW', root.quit)
    root.mainloop()
    chosen = [k for k, _, _ in CATEGORIES if vars_[k].get()] if result['ok'] else None
    root.destroy()
    return chosen


def pick_save_path():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
        os.makedirs(OUT_DIR, exist_ok=True)
        path = filedialog.asksaveasfilename(title='Куда сохранить выгрузку?', initialdir=OUT_DIR,
                                            initialfile=os.path.basename(OUT), defaultextension='.xlsx',
                                            filetypes=[('Excel', '*.xlsx')])
        root.destroy()
        return path or None
    except Exception:
        return OUT


def build_workbook(data, chosen_keys):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    bold = Font(bold=True)
    price_fill = PatternFill('solid', fgColor='FFF6D5')
    band_fill = PatternFill('solid', fgColor='E4E8EE')
    # Справка
    ws = wb.active; ws.title = 'Справка'
    for line in HELP_TEXT:
        ws.append([line])
    ws.column_dimensions['A'].width = 105
    ws['A1'].font = bold
    # Категории — единые колонки на всех листах
    for k, _, builder in CATEGORIES:
        if k not in chosen_keys:
            continue
        spec = builder(data)
        ws = wb.create_sheet(spec['title'])
        ws.append(HEADERS)
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=1, column=c).font = bold
        for i, wdt in enumerate(WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        ws.freeze_panes = 'B2'
        for r in spec['rows']:
            ws.append(r)
            if str(r[0] or '').startswith('#'):   # заголовок таблицы внутри листа
                ws.merge_cells(start_row=ws.max_row, start_column=1,
                               end_row=ws.max_row, end_column=len(HEADERS))
                cell = ws.cell(row=ws.max_row, column=1)   # «#» в начале — по нему загрузка пропускает
                cell.font = bold
                cell.fill = band_fill
                cell.alignment = Alignment(vertical='center')
                continue
            ws.cell(row=ws.max_row, column=PRICE_COL).fill = price_fill
            # «Цвет (hex)» (колонка 11) — заливаем ячейку САМИМ цветом, чтобы видеть его глазами
            # (задание: «чтоб я видел визуально цвет»). Для текстур цвета обычно нет — ячейка пустая.
            hexv = str(r[HEX_COL - 1] or '')
            if hexv.startswith('#') and len(hexv) == 7:
                ws.cell(row=ws.max_row, column=HEX_COL).fill = PatternFill('solid', fgColor='FF' + hexv[1:].upper())
        add_delete_dropdown(ws)   # «Удалить» — галочка-выбор на каждом листе
        if k == 'ldsp':      # да/нет по поверхностям есть только у материалов
            add_yes_no_dropdown(ws)
        if k == 'hardware':  # назначение добавляемой фурнитуры — выбором, не руками
            add_purpose_dropdown(ws)
    return wb


def add_purpose_dropdown(ws):
    """Выпадашка назначений в «От чего зависит» — на нижнюю (ДОБАВЛЯЕМУЮ) часть листа фурнитуры."""
    from openpyxl.worksheet.datavalidation import DataValidation
    start = next((r for r in range(1, ws.max_row + 1)
                  if str(ws.cell(row=r, column=1).value or '').startswith('# ДОБАВЛЯЕМЫЕ')), None)
    if start is None:
        return
    dv = DataValidation(type='list', formula1='"' + ','.join(FIT_PURPOSES) + '"',
                        allow_blank=True, showDropDown=False)
    dv.errorTitle = 'Назначение фурнитуры'
    dv.error = ('Выберите назначение из списка. Другое сюда писать нельзя: конфигуратор умеет '
                'выбирать только эти позиции, остальное просто осело бы в каталоге.')
    dv.promptTitle = 'Назначение'
    dv.prompt = 'Куда эта фурнитура ставится (по нему конфигуратор её и предложит)'
    ws.add_data_validation(dv)
    dv.add(f'K{start + 1}:K{ws.max_row + YES_NO_SPARE_ROWS}')   # K = «От чего зависит» (сдвиг из-за «Удалить»)


# «Корпус/Фасад/Наполнение» — выбор из списка да/нет прямо в ячейке (просьба пользователя 6.08):
# щёлкнул — выбрал, руками не пишешь и не ошибаешься в написании. Запас строк вниз — чтобы
# выпадашка работала и в НОВЫХ строках, которые пользователь добавит сам.
YES_NO_COLS = (13, 14, 15)   # Корпус | Фасад | Наполнение
HEX_COL = 12                 # «Цвет (hex)» — ячейка заливается самим цветом
DELETE_COL = 1               # «Удалить» — крайний слева, «да» = убрать позицию (7.08)
YES_NO_SPARE_ROWS = 300


def add_delete_dropdown(ws):
    """Крайний слева столбец «Удалить»: выбор «да» = убрать позицию из каталога (задание 7.08 —
    «галочку понятнее, чем слово в цене»). Пусто = позицию не трогаем."""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    dv = DataValidation(type='list', formula1='"да"', allow_blank=True, showDropDown=False)
    dv.errorTitle = 'Удаление позиции'
    dv.error = 'Здесь только «да» — выберите из списка (пусто = позицию не трогаем).'
    dv.promptTitle = 'Удалить позицию'
    dv.prompt = 'да — позиция будет удалена из каталога при загрузке'
    ws.add_data_validation(dv)
    col = get_column_letter(DELETE_COL)
    dv.add(f'{col}2:{col}{ws.max_row + YES_NO_SPARE_ROWS}')


def add_yes_no_dropdown(ws):
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    last = ws.max_row + YES_NO_SPARE_ROWS
    dv = DataValidation(type='list', formula1='"да,нет"', allow_blank=True, showDropDown=False)
    dv.error = 'Здесь только «да» или «нет» — выберите из списка.'
    dv.errorTitle = 'Доступно ли в этой поверхности'
    dv.prompt = 'да — материал доступен здесь; нет — недоступен'
    dv.promptTitle = 'да / нет'
    ws.add_data_validation(dv)
    for c in YES_NO_COLS:
        col = get_column_letter(c)
        dv.add(f'{col}2:{col}{last}')


def main():
    global OUT, OUT_DIR
    args = sys.argv[1:]
    if args and args[0] == '--all':
        chosen = [k for k, _, _ in CATEGORIES]
        OUT = args[1] if len(args) > 1 else OUT
    else:
        # --dir <папка>: куда по умолчанию сохранять (батник передаёт Config\Выгрузки)
        if len(args) >= 2 and args[0] == '--dir':
            OUT_DIR = args[1]
            OUT = os.path.join(OUT_DIR, 'ассортимент.xlsx')
        chosen = choose_categories()
        if not chosen:
            print('Выгрузка отменена.'); return 1
        picked = pick_save_path()
        if not picked:
            print('Выгрузка отменена — путь не выбран.'); return 1
        OUT = picked
    data = load()
    wb = build_workbook(data, chosen)
    os.makedirs(os.path.dirname(os.path.abspath(OUT)) or '.', exist_ok=True)
    try:
        wb.save(OUT)
    except PermissionError:
        print(f'ОШИБКА: файл «{os.path.basename(OUT)}» открыт в Excel — закройте и повторите.'); return 1
    print(f'Готово: {OUT}')
    print('Листы: ' + ', '.join(wb.sheetnames))
    return 0


if __name__ == '__main__':
    code = main()
    if not (len(sys.argv) > 1 and sys.argv[1] == '--all'):
        input('\nНажмите Enter, чтобы закрыть...')
    sys.exit(code)
