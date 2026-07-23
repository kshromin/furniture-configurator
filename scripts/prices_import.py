# -*- coding: utf-8 -*-
# Загрузка цен и ассортимента из «для работы/цены.xlsx» обратно в data/materials.json
# (задание «скрипт для цен 18,07», пара к prices_export.py).
#
# Принцип: сначала ВСЕ проверки — при любой ошибке ничего не меняется, печатается список
# ошибок словами. Только если ошибок нет: бэкап старого materials.json в
# «для работы/бэкапы цен/», запись нового, отчёт что изменилось.
#
# Правила:
#  - существующие позиции находятся по скрытым _id/_key; менять их руками нельзя;
#  - удалять строки нельзя (вывод из ассортимента — отдельный будущий скрипт);
#  - листы «только цены» (Направляющие, Сетчатые полки, Корзины, Фурнитура) новых строк
#    не принимают; листы ассортимента (ЛДСП, Профили, Цвета профилей, Стёкла, Услуги) —
#    строка без _id = новая позиция, id генерируется автоматически;
#  - все прочие поля json (единицы измерения, служебные имена) не трогаются.
import json
import os
import re
import shutil
import sys
import time
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'data', 'materials.json')
# Папки систематизированы (21.07): выгрузки — Config/Выгрузки, бэкапы — Config/Бэкапы
XLSX = os.path.normpath(os.path.join(ROOT, '..', 'Выгрузки', 'цены.xlsx'))
BACKUP_DIR = os.path.normpath(os.path.join(ROOT, '..', 'Бэкапы'))

LDSP_GROUPS = {'ЛДСП корпус': 'korpus', 'ЛДСП фасад': 'fasad', 'ЛДСП наполнение': 'fill'}

errors = []
changes = []
_id_counter = [0]


def new_id(prefix):
    _id_counter[0] += 1
    return f'{prefix}{int(time.time())}_{_id_counter[0]}'


def cell_str(v):
    return str(v).strip() if v is not None else ''


def parse_price(v, where, allow_fraction=False):
    if v is None or cell_str(v) == '':
        errors.append(f'{where}: цена не заполнена')
        return None
    try:
        n = float(str(v).replace(',', '.').replace(' ', ''))
    except ValueError:
        errors.append(f'{where}: цена «{v}» — не число')
        return None
    if n <= 0:
        errors.append(f'{where}: цена должна быть больше нуля (сейчас {n})')
        return None
    if not allow_fraction and abs(n - round(n)) < 1e-9:
        n = round(n)
    return n


def parse_hex(v, where, required=True):
    s = cell_str(v)
    if not s:
        if required:
            errors.append(f'{where}: не заполнен цвет (hex)')
        return None
    s = s if s.startswith('#') else '#' + s
    if not re.fullmatch(r'#[0-9a-fA-F]{6}', s):
        errors.append(f'{where}: цвет «{v}» — не hex вида #f4f3f0')
        return None
    return s.lower()


def rows_of(wb, title, ncols):
    if title not in wb.sheetnames:
        errors.append(f'Лист «{title}» не найден в файле')
        return []
    out = []
    for r, row in enumerate(wb[title].iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row[:ncols]) + [None] * (ncols - len(row))
        if all(cell_str(v) == '' for v in vals):
            continue
        out.append((r, vals))
    return out


def set_price(obj, field, val, label):
    if obj[field] != val:
        changes.append(f'{label}: {obj[field]} → {val}')
        obj[field] = val


def pick_xlsx():
    """Окно выбора файла (просьба 21.07): по умолчанию «цены.xlsx», но можно выбрать любой
    (например, файл, присланный по почте). Если окно недоступно — берётся цены.xlsx."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        path = filedialog.askopenfilename(
            title='Какой файл с ценами загрузить?',
            initialdir=os.path.dirname(XLSX),
            initialfile=os.path.basename(XLSX),
            filetypes=[('Excel', '*.xlsx')])
        root.destroy()
        return path or None  # закрыл окно без выбора = отмена
    except Exception:
        return XLSX if os.path.exists(XLSX) else None


def main():
    from openpyxl import load_workbook

    # Путь можно передать аргументом (без окна выбора) — для автоматизации/проверок
    xlsx = sys.argv[1] if len(sys.argv) > 1 else pick_xlsx()
    if not xlsx:
        print('Загрузка отменена — файл не выбран.')
        return 1
    if not os.path.exists(xlsx):
        print(f'Файл не найден: {xlsx}\nСначала запустите «Выгрузить цены.bat».')
        return 1
    print(f'Загружаю: {xlsx}\n')
    with open(SRC, encoding='utf-8') as f:
        data = json.load(f)
    wb = load_workbook(xlsx, data_only=True)

    # ── Палитра (служебный справочник «имя → hex») — пересобирается из листа целиком;
    # по ней остальные листы разрешают цвет, введённый по имени ──
    palette = []
    pal_names = set()
    for r, (name, hexv) in rows_of(wb, 'Палитра', 2):
        where = f'«Палитра», строка {r}'
        name = cell_str(name)
        hx = parse_hex(hexv, where)
        if not name:
            errors.append(f'{where}: не заполнено название')
            continue
        if hx is None:
            continue
        if name.lower() in pal_names:
            errors.append(f'{where}: цвет «{name}» в палитре дважды')
            continue
        pal_names.add(name.lower())
        palette.append({'name': name, 'hex': hx})
    pal_by_name = {p['name'].lower(): p['hex'] for p in palette}
    if palette and data.get('palette') != palette:
        if json.dumps(data.get('palette', []), sort_keys=True) != json.dumps(palette, sort_keys=True):
            changes.append(f'Палитра обновлена ({len(palette)} цветов)')
        data['palette'] = palette

    def resolve_color(v, where):
        """Цвет из ячейки: имя из палитры или сразу hex."""
        s = cell_str(v)
        if not s:
            errors.append(f'{where}: не заполнен цвет')
            return None
        if re.fullmatch(r'#?[0-9a-fA-F]{6}', s):
            return (s if s.startswith('#') else '#' + s).lower()
        hx = pal_by_name.get(s.lower())
        if not hx:
            errors.append(f'{where}: цвета «{s}» нет в палитре — добавьте его на лист «Палитра» '
                          f'или впишите hex вида #f4f3f0')
        return hx

    # ── ЛДСП (3 листа): правки по _id, новые строки = новые цвета ──
    for title, group in LDSP_GROUPS.items():
        producers = data[group]['producers']
        by_id = {c['id']: c for p in producers for c in p['colors']}
        prod_by_name = {p['name'].strip().lower(): p for p in producers}
        seen = set()
        # Цвет (hex) у ЛДСП не задаётся (просьба 21.07): вид даст текстура, без неё конфигуратор
        # покажет коричневую заглушку. Старый hex существующих цветов не трогается.
        for r, (prod_name, name, price, texture, cid, _pid) in rows_of(wb, title, 6):
            where = f'«{title}», строка {r}'
            if cell_str(cid) in by_id:
                seen.add(cell_str(cid))  # строка на месте, даже если в ней ошибка — без каскада «удалено»
            price = parse_price(price, where)
            name = cell_str(name)
            texture = cell_str(texture)
            if not name:
                errors.append(f'{where}: не заполнено название цвета')
            if price is None or not name:
                continue
            cid = cell_str(cid)
            if cid:
                if cid not in by_id:
                    errors.append(f'{where}: служебный id «{cid}» не найден — колонку _id менять нельзя')
                    continue
                seen.add(cid)
                col = by_id[cid]
                set_price(col, 'pricePerM2', price, f'{title} / {name}')
                if col['name'] != name:
                    changes.append(f'{title}: «{col["name"]}» переименован в «{name}»')
                    col['name'] = name
                if texture:
                    if col.get('texture') != texture:
                        changes.append(f'{title} / {name}: текстура → {texture}')
                        col['texture'] = texture
                elif 'texture' in col:
                    changes.append(f'{title} / {name}: текстура убрана')
                    del col['texture']
            else:
                prod = prod_by_name.get(cell_str(prod_name).lower())
                if not prod:
                    # Новый производитель создаётся сам (просьба 21.07) — по имени в строке
                    pn = cell_str(prod_name)
                    if not pn:
                        errors.append(f'{where}: не заполнен производитель')
                        continue
                    prod = {'id': new_id(group[0] + 'p'), 'name': pn, 'colors': []}
                    producers.append(prod)
                    prod_by_name[pn.lower()] = prod
                    changes.append(f'{title}: создан производитель «{pn}»')
                col = {'id': new_id(group[0]), 'name': name, 'pricePerM2': price}
                if texture:
                    col['texture'] = texture
                prod['colors'].append(col)
                changes.append(f'{title}: добавлен цвет «{name}» ({prod["name"]}, {price} ₽/м², '
                               f'вид — текстура либо коричневая заглушка)')
        missing = set(by_id) - seen
        if missing:
            errors.append(f'«{title}»: удалены строки ({", ".join(sorted(missing))}) — '
                          f'удаление через этот файл запрещено')

    # ── Цвета профилей (ассортимент; цены — на листе «Профили купе») ──
    colors = data['slidingDoor']['colors']
    by_id = {c['id']: c for c in colors}
    seen = set()
    for r, (name, hexv, cid) in rows_of(wb, 'Цвета профилей', 3):
        where = f'«Цвета профилей», строка {r}'
        if cell_str(cid) in by_id:
            seen.add(cell_str(cid))
        name = cell_str(name)
        hx = resolve_color(hexv, where)
        if not name:
            errors.append(f'{where}: не заполнено название')
        if hx is None or not name:
            continue
        cid = cell_str(cid)
        if cid:
            if cid not in by_id:
                errors.append(f'{where}: служебный id «{cid}» не найден')
                continue
            c = by_id[cid]
            if c['hex'] != hx:
                changes.append(f'Цвет профиля {name}: {c["hex"]} → {hx}')
                c['hex'] = hx
            if c['name'] != name:
                changes.append(f'Цвет профиля «{c["name"]}» переименован в «{name}»')
                c['name'] = name
        else:
            colors.append({'id': new_id('pcol'), 'name': name, 'hex': hx})
            changes.append(f'Цвета профилей: добавлен «{name}» (заполните его цены на листе «Профили купе»)')
    if set(by_id) - seen:
        errors.append('«Цвета профилей»: часть строк удалена — удаление запрещено')

    # ── Профили купе: полный прайс по сочетаниям профиль×цвет ──
    profiles = data['slidingDoor']['profiles']
    price_list = data['slidingDoor'].setdefault('profilePrices', [])
    by_key = {f"{x['profile']}:{x['color']}": x for x in price_list}
    prof_by_name = {p['name'].strip().lower(): p for p in profiles}
    col_by_name = {c['name'].strip().lower(): c for c in colors}
    seen = set()
    for r, (pname, cname, vert, top, bottom, key) in rows_of(wb, 'Профили купе', 6):
        where = f'«Профили купе», строка {r}'
        if cell_str(key) in by_key:
            seen.add(cell_str(key))
        vert, top, bottom = (parse_price(v, where) for v in (vert, top, bottom))
        if None in (vert, top, bottom):
            continue
        key = cell_str(key)
        if key:
            if key not in by_key:
                errors.append(f'{where}: служебный ключ «{key}» не найден — колонку _key менять нельзя')
                continue
            x = by_key[key]
            label = f'Профиль {cell_str(pname)} × {cell_str(cname)}'
            set_price(x, 'vertPerM', vert, f'{label}, вертикаль')
            set_price(x, 'horizTopPerM', top, f'{label}, гориз. верх')
            set_price(x, 'horizBottomPerM', bottom, f'{label}, гориз. низ')
        else:
            # Новая строка прайса: цвет должен существовать, профиль создаётся сам по имени
            col = col_by_name.get(cell_str(cname).lower())
            if not col:
                errors.append(f'{where}: цвет «{cname}» не найден — сначала добавьте его '
                              f'на листе «Цвета профилей»')
                continue
            pn = cell_str(pname)
            if not pn:
                errors.append(f'{where}: не заполнен профиль')
                continue
            prof = prof_by_name.get(pn.lower())
            if not prof:
                prof = {'id': new_id('prof'), 'name': pn, 'vertPerM': vert,
                        'horizTopPerM': top, 'horizBottomPerM': bottom}
                profiles.append(prof)
                prof_by_name[pn.lower()] = prof
                changes.append(f'Профили купе: создан профиль «{pn}» (рамка в 3D — стандартной ширины)')
            k = f"{prof['id']}:{col['id']}"
            if k in by_key:
                errors.append(f'{where}: сочетание {pn} × {col["name"]} уже есть в прайсе выше')
                continue
            x = {'profile': prof['id'], 'color': col['id'],
                 'vertPerM': vert, 'horizTopPerM': top, 'horizBottomPerM': bottom}
            price_list.append(x)
            by_key[k] = x
            seen.add(k)
            changes.append(f'Профили купе: добавлен прайс {pn} × {col["name"]}')
    missing_rows = set(by_key) - seen
    if missing_rows:
        errors.append('«Профили купе»: часть строк удалена — удаление запрещено')
    # Полнота прайса: у каждой пары профиль×цвет должна быть цена — иначе конфигуратор
    # не сможет посчитать это сочетание
    for p in profiles:
        for c in colors:
            if f"{p['id']}:{c['id']}" not in by_key:
                errors.append(f'«Профили купе»: нет цен для сочетания {p["name"]} × {c["name"]} — '
                              f'добавьте строку')

    # ── Наполнение дверей: зеркало (цена) + стёкла (ассортимент) ──
    fills = data['slidingDoor']['fills']
    glass = fills.setdefault('glass', {'name': 'Стекло', 'colors': []})
    by_id = {c['id']: c for c in glass['colors']}
    seen = set()
    for r, (typ, name, hexv, price, cid) in rows_of(wb, 'Наполнение дверей', 5):
        where = f'«Наполнение дверей», строка {r}'
        if cell_str(cid) in by_id:
            seen.add(cell_str(cid))
        price = parse_price(price, where)
        if price is None:
            continue
        cid = cell_str(cid)
        if cid == 'mirror':
            set_price(fills['mirror'], 'pricePerM2', price, 'Зеркало')
            continue
        name = cell_str(name)
        hx = resolve_color(hexv, where)
        if not name:
            errors.append(f'{where}: не заполнено название')
        if hx is None or not name:
            continue
        if cid:
            if cid not in by_id:
                errors.append(f'{where}: служебный id «{cid}» не найден')
                continue
            seen.add(cid)
            c = by_id[cid]
            set_price(c, 'pricePerM2', price, f'Стекло {name}')
            if c['color'] != hx:
                changes.append(f'Стекло {name}: цвет {c["color"]} → {hx}')
                c['color'] = hx
            if c['name'] != name:
                changes.append(f'Стекло «{c["name"]}» переименовано в «{name}»')
                c['name'] = name
        else:
            glass['colors'].append({'id': new_id('gl'), 'name': name, 'color': hx, 'pricePerM2': price})
            changes.append(f'Наполнение дверей: добавлено стекло «{name}» ({price} ₽/м²)')
    if set(by_id) - seen:
        errors.append('«Наполнение дверей»: часть строк стёкол удалена — удаление запрещено')

    # ── Только цены: направляющие / сетчатые полки / корзины ──
    def price_only(title, items, key_fn, label_fn, field, key_col, price_col, ncols):
        by_key = {key_fn(it): it for it in items}
        for r, vals in rows_of(wb, title, ncols):
            where = f'«{title}», строка {r}'
            key = cell_str(vals[key_col])
            if not key:
                # серые строки-разделители блоков (без ключа и без цены) — просто пропускаем
                if cell_str(vals[price_col]) == '':
                    continue
                errors.append(f'{where}: новая строка — на этом листе можно менять только цены')
                continue
            if key not in by_key:
                errors.append(f'{where}: служебный ключ «{key}» не найден — колонку _key менять нельзя')
                continue
            price = parse_price(vals[price_col], where)
            if price is not None:
                set_price(by_key[key], field, price, label_fn(by_key[key]))

    price_only('Направляющие', data['drawerSlide'], lambda s: f"{s['type']}:{s['length']}",
               lambda s: f"Направляющие {s['type']} {s['length']}мм", 'price', 3, 2, 4)

    # ── Сетчатые полки и Корзины — размерные сетки теперь живут в каталоге (разделение
    # «как строим»/«из чего строим», 21.07): новая строка без ключа = новый размер/цвет,
    # конфигуратор подхватывает сетки из данных сам. Цвета — фиксированный набор (3D-материалы).
    METAL_COLORS = {'белый': 'white', 'серебро': 'silver', 'чёрный': 'black', 'черный': 'black',
                    'white': 'white', 'silver': 'silver', 'black': 'black'}

    def parse_metal_color(v, where):
        c = METAL_COLORS.get(cell_str(v).lower())
        if not c:
            errors.append(f'{where}: цвет «{v}» — допустимы Белый, Серебро, Чёрный')
        return c

    def parse_size(v, where, label):
        try:
            n = int(float(str(v).replace(',', '.')))
            if n <= 0:
                raise ValueError
            return n
        except (ValueError, TypeError):
            errors.append(f'{where}: {label} «{v}» — не целое число миллиметров')
            return None

    mesh_by_key = {f"{m['depth']}:{m['color']}": m for m in data['meshShelf']}
    for r, (name, depth, colorv, price, key) in rows_of(wb, 'Сетчатые полки', 5):
        where = f'«Сетчатые полки», строка {r}'
        price = parse_price(price, where)
        key = cell_str(key)
        if key:
            if key not in mesh_by_key:
                errors.append(f'{where}: служебный ключ «{key}» не найден — колонку _key менять нельзя')
                continue
            m = mesh_by_key[key]
            if price is not None:
                set_price(m, 'pricePerM', price, f'Сетчатая полка {m["name"]}')
            nm = cell_str(name)
            if nm and m['name'] != nm:
                changes.append(f'Сетчатая полка «{m["name"]}» переименована в «{nm}»')
                m['name'] = nm
        else:
            depth = parse_size(depth, where, 'глубина')
            color = parse_metal_color(colorv, where)
            nm = cell_str(name)
            if not nm:
                errors.append(f'{where}: не заполнено название')
            if None in (depth, color, price) or not nm:
                continue
            k = f'{depth}:{color}'
            if k in mesh_by_key:
                errors.append(f'{where}: сетчатая полка {depth}мм/{cell_str(colorv)} уже есть выше')
                continue
            m = {'depth': depth, 'color': color, 'name': nm, 'pricePerM': price}
            data['meshShelf'].append(m)
            mesh_by_key[k] = m
            changes.append(f'Сетчатые полки: добавлена «{nm}» ({depth}мм, {price} ₽/пог.м)')

    basket_by_key = {f"{b['width']}:{b['depth']}:{b['height']}:{b['color']}": b for b in data['basket']}
    for r, (w, dp, h, colorv, price, key) in rows_of(wb, 'Корзины', 6):
        where = f'«Корзины», строка {r}'
        price = parse_price(price, where)
        key = cell_str(key)
        if key:
            if key not in basket_by_key:
                errors.append(f'{where}: служебный ключ «{key}» не найден — колонку _key менять нельзя')
                continue
            b = basket_by_key[key]
            if price is not None:
                set_price(b, 'price', price, f'Корзина {b["width"]}×{b["depth"]}×{b["height"]} {b["color"]}')
        else:
            w = parse_size(w, where, 'ширина')
            dp = parse_size(dp, where, 'глубина')
            h = parse_size(h, where, 'высота')
            color = parse_metal_color(colorv, where)
            if None in (w, dp, h, color, price):
                continue
            k = f'{w}:{dp}:{h}:{color}'
            if k in basket_by_key:
                errors.append(f'{where}: корзина {w}×{dp}×{h} {cell_str(colorv)} уже есть выше')
                continue
            b = {'width': w, 'depth': dp, 'height': h, 'color': color, 'price': price}
            data['basket'].append(b)
            basket_by_key[k] = b
            changes.append(f'Корзины: добавлена {w}×{dp}×{h} {cell_str(colorv)} ({price} ₽)')

    # ── Фурнитура и разное (общий лист, ключ-путь) ──
    sd = data['slidingDoor']
    flat = {f'fittings:{it["id"]}': (it, 'price', it['name']) for it in data['fittings']}
    flat['swing'] = (data['swingDoorHardware'], 'pricePerDoor', 'Петли распашных')
    flat['rollers'] = (sd['rollers'], 'pricePerSet', 'Ролики купе')
    flat['track'] = (sd['track'], 'pricePerM', 'Направляющая купе')
    flat['divider'] = (sd['divider'], 'pricePerM', 'Перемычка купе')
    flat['edge'] = (data['edgeBanding'], 'pricePerM', 'Кромка')
    for r, (name, price, key) in rows_of(wb, 'Фурнитура и разное', 3):
        where = f'«Фурнитура и разное», строка {r}'
        key = cell_str(key)
        if not key and cell_str(price) == '':
            continue  # серая строка-разделитель блока единиц измерения
        if key not in flat:
            errors.append(f'{where}: новая строка — на этом листе можно менять только цены')
            continue
        price = parse_price(price, where)
        if price is not None:
            obj, field, label = flat[key]
            set_price(obj, field, price, label)

    # ── Услуги (extras) — расширяемые в рамках существующих групп ──
    groups_by_id = {g['id']: g for g in data['extras']}
    groups_by_name = {g['name'].strip().lower(): g for g in data['extras']}
    by_id = {it['id']: (g, it) for g in data['extras'] for it in g['items']}
    seen = set()
    for r, (grp_name, name, price, gid, iid) in rows_of(wb, 'Услуги', 5):
        where = f'«Услуги», строка {r}'
        if cell_str(iid) in by_id:
            seen.add(cell_str(iid))
        name = cell_str(name)
        price = parse_price(price, where)
        if not name:
            errors.append(f'{where}: не заполнено название')
        if price is None or not name:
            continue
        iid = cell_str(iid)
        if iid:
            if iid not in by_id:
                errors.append(f'{where}: служебный id «{iid}» не найден')
                continue
            seen.add(iid)
            _, it = by_id[iid]
            set_price(it, 'price', price, f'Услуга «{name}»')
            if it['name'] != name:
                changes.append(f'Услуга «{it["name"]}» переименована в «{name}»')
                it['name'] = name
        else:
            grp = groups_by_id.get(cell_str(gid)) or groups_by_name.get(cell_str(grp_name).lower())
            if not grp:
                errors.append(f'{where}: группа «{grp_name}» не найдена '
                              f'(есть: {", ".join(g["name"] for g in data["extras"])})')
                continue
            grp['items'].append({'id': new_id('ex'), 'name': name, 'price': price})
            changes.append(f'Услуги / {grp["name"]}: добавлена «{name}» ({price} ₽)')
    if set(by_id) - seen:
        errors.append('«Услуги»: часть строк удалена — удаление запрещено')

    # ── Итог ──
    if errors:
        print(f'НИЧЕГО НЕ ЗАГРУЖЕНО — найдено ошибок: {len(errors)}\n')
        for e in errors:
            print(' •', e)
        print(f'\nИсправьте файл «{os.path.basename(xlsx)}» и запустите загрузку ещё раз.')
        return 1

    if not changes:
        print('Изменений не найдено — файл совпадает с текущим каталогом, ничего не менялось.')
        return 0

    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup = os.path.join(BACKUP_DIR, f'materials-{datetime.now():%Y%m%d-%H%M%S}.json')
    shutil.copy2(SRC, backup)
    with open(SRC, 'w', encoding='utf-8', newline='\n') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')

    print(f'Загружено изменений: {len(changes)}\n')
    for c in changes:
        print(' •', c)
    print(f'\nСтарый каталог сохранён: {backup}')
    print('Обновите страницу конфигуратора, чтобы увидеть новые цены.')
    return 0


if __name__ == '__main__':
    code = main()
    input('\nНажмите Enter, чтобы закрыть...')
    sys.exit(code)
